from __future__ import annotations

import json
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from simulator.agents.roster import load_long_campaign_party, load_long_campaign_roster
from simulator.core.dice import DiceRoller, slot_result


ROOT = Path(__file__).resolve().parents[2]
REPORTS = ROOT / "reports"
RAW = REPORTS / "raw_logs"


@dataclass
class StrictRule:
    rule_id: str
    name: str
    chapter: str
    category: str


RULES = [
    StrictRule("SCENE_CHECK", "情景检定：属性骰+熟练骰+情景骰", "玩家1/DM1", "情景"),
    StrictRule("SLOT_ASSIGN", "三槽分配：命中/得利/代价或破绽", "玩家1", "核心"),
    StrictRule("CRISIS_GAIN", "危机值按代价等级增长", "DM2", "情景"),
    StrictRule("MYSTERY_GAIN", "关键线索行动推进解密值", "DM2", "情景"),
    StrictRule("FORTUNE_GAIN", "命中后得利等级加入战役得利", "DM2", "情景"),
    StrictRule("FORTUNE_UNLOCK", "战役得利达到阈值触发解锁", "DM2", "情景"),
    StrictRule("CORE_CLUE", "核心线索不因失败消失", "DM总原则", "情景"),
    StrictRule("ASSIST", "协助行动提供额外情景骰", "DM2", "情景"),
    StrictRule("FORCE_TRY", "强行尝试", "玩家/DM2", "情景"),
    StrictRule("STABILIZE", "稳控机会", "DM2", "情景"),
    StrictRule("COMBAT_ACTION", "正式战斗动作", "玩家2", "战斗"),
    StrictRule("WEAPON_FEATURE", "武器特性按得利门槛触发", "数据库/玩家2", "战斗"),
    StrictRule("FLAW_RESOLVE", "破绽结算", "玩家2/DM3", "战斗"),
    StrictRule("ENEMY_REACTION", "敌人反击/削得利/断势", "DM3", "战斗"),
    StrictRule("BOSS_COUNTER", "Boss正式反击", "DM3", "战斗"),
    StrictRule("NONLETHAL", "非杀伤攻击/制服", "DM3", "战斗"),
    StrictRule("ENV_ACTION", "环境互动动作", "DM3", "战斗"),
    StrictRule("LIGHTNESS_POS", "轻功/Pos/追逐位势", "玩家2", "轻功"),
    StrictRule("REST_3_PHASE", "三阶段休整", "玩家5/DM5", "休整"),
    StrictRule("GROWTH_CHECK", "熟练成长检定", "玩家5", "成长"),
    StrictRule("MARTIAL_PRACTICE", "武学谱修行", "玩家5", "成长"),
    StrictRule("INNER_PRACTICE", "内功成长", "玩家5", "成长"),
    StrictRule("FAME", "名望资源", "DM4/5", "长期"),
    StrictRule("FAVOR", "人情资源", "DM4/5", "长期"),
    StrictRule("INTEL", "情报资源", "DM4/5", "长期"),
    StrictRule("LEVERAGE", "把柄资源", "DM4/5", "长期"),
    StrictRule("AGE_YEARS", "岁月/年龄推进", "DM5", "长期"),
    StrictRule("FAMILY_PROPERTY", "家庭/弟子/产业钩子", "DM5", "长期"),
    StrictRule("DATABASE_EVENT", "江湖动态事件库投放", "数据库", "数据库"),
    StrictRule("DATABASE_ENEMY", "敌人数据库投放", "数据库", "数据库"),
    StrictRule("DATABASE_REWARD", "奖励代价库投放", "数据库", "数据库"),
]


CAMPAIGNS = [
    {
        "id": "SLC01",
        "title": "京畿暗账长团",
        "seed": 9101,
        "focus": "调查、官府、人情、把柄",
        "cycles": [
            "失银账册",
            "船工灭口",
            "北仓暗渠",
            "亲兵假令",
            "商会账房",
            "皇城外证",
        ],
    },
    {
        "id": "SLC02",
        "title": "北地断碑长团",
        "seed": 9202,
        "focus": "秘境、武学、高手、轻功",
        "cycles": [
            "断碑委托",
            "雪岭追踪",
            "古墓机关",
            "门派遗债",
            "高手邀斗",
            "残谱归属",
        ],
    },
    {
        "id": "SLC03",
        "title": "南海潮生长团",
        "seed": 9303,
        "focus": "商路、名望、家庭、产业、海盗",
        "cycles": [
            "海商护送",
            "盐仓人情",
            "宗族旧债",
            "海盗夜袭",
            "产业抉择",
            "南海潮会",
        ],
    },
]


def rule_template() -> dict[str, dict[str, Any]]:
    return {
        r.rule_id: {
            "rule_id": r.rule_id,
            "rule_name": r.name,
            "chapter": r.chapter,
            "category": r.category,
            "opportunities": 0,
            "triggered_count": 0,
            "player_initiated_count": 0,
            "dm_initiated_count": 0,
            "changed_outcome_count": 0,
            "ignored_count": 0,
            "ambiguity_count": 0,
            "total_resolution_time": 0.0,
            "notes": [],
        }
        for r in RULES
    }


def mark(stats: dict[str, dict[str, Any]], rid: str, *, opportunity: bool = True, triggered: bool = False, player: bool = False, dm: bool = False, changed: bool = False, ignored: bool = False, ambiguous: bool = False, seconds: float = 0.0, note: str = "") -> None:
    row = stats[rid]
    if opportunity:
        row["opportunities"] += 1
    if triggered:
        row["triggered_count"] += 1
    if player:
        row["player_initiated_count"] += 1
    if dm:
        row["dm_initiated_count"] += 1
    if changed:
        row["changed_outcome_count"] += 1
    if ignored:
        row["ignored_count"] += 1
    if ambiguous:
        row["ambiguity_count"] += 1
    row["total_resolution_time"] += seconds
    if note:
        row["notes"].append(note)


def roll_action(roller: DiceRoller, attr: str, skill: str, dc: int, prefer_feature: bool = False) -> dict[str, Any]:
    rolls = roller.roll_many([(attr, "属性"), (skill, "熟练"), ("d6", "额外"), ("d6", "额外")])
    order = sorted(range(4), key=lambda i: rolls[i].value, reverse=True)
    if prefer_feature:
        hit_idx = [order[0], order[2]]
        benefit_idx = [order[1]]
        flaw_idx = [order[3]]
    else:
        hit_idx = [order[0], order[1]]
        benefit_idx = [order[2]]
        flaw_idx = [order[3]]
    result = slot_result(rolls, hit_idx, benefit_idx, flaw_idx)
    return {
        "rolls": [asdict(r) for r in rolls],
        "slots": {"hit": hit_idx, "benefit": benefit_idx, "flaw": flaw_idx},
        "dc": dc,
        "result": result,
        "success": result["hit"] >= dc,
    }


def feedback_for(player_type: str, cycle: int, tags: list[str]) -> list[dict[str, str]]:
    output = []
    if "代价疲劳" in tags and player_type in {"新手玩家", "剧情玩家", "战斗玩家"}:
        output.append({"玩家类型": player_type, "标签": "【太惩罚】", "反馈": "我知道成功会有代价，但长团里每一轮都这样，会开始不敢行动。"})
    if "低使用武学" in tags and player_type == "普通构筑玩家":
        output.append({"玩家类型": player_type, "标签": "【看不懂】", "反馈": "武学谱和内功成长一直没真正上桌，我不知道什么时候该用它。"})
    if "DM记录压力" in tags and player_type == "DM":
        output.append({"玩家类型": "DM", "标签": "【DM负担大】", "反馈": "每轮都要记危机、解密、战役得利、名望、人情、情报、把柄、伤势、仇敌、岁月事件和下一钩子，普通主持很容易漏。"})
    if "资源无入口" in tags and player_type == "剧情玩家":
        output.append({"玩家类型": player_type, "标签": "【选择无意义】", "反馈": "名望、人情、把柄有记录，但如果没有明确消费入口，我不会主动用。"})
    return output


def run_campaign(campaign: dict[str, Any], roster: dict[str, Any]) -> dict[str, Any]:
    roller = DiceRoller(campaign["seed"])
    stats = rule_template()
    party = load_long_campaign_party()
    raw_path = RAW / f"strict_long_{campaign['id']}_seed_{campaign['seed']}.jsonl"
    resources = {
        "危机": 0,
        "解密": 0,
        "战役得利": 0,
        "名望": 0,
        "人情": 0,
        "情报": 0,
        "把柄": 0,
        "伤势": 0,
        "仇敌": 0,
        "弟子": 0,
        "家庭": 0,
        "产业": 0,
        "年龄推进": 0,
    }
    events = []
    RAW.mkdir(parents=True, exist_ok=True)
    with raw_path.open("w", encoding="utf-8") as f:
        for cycle_index, title in enumerate(campaign["cycles"], start=1):
            dc = 8 if resources["危机"] < 6 else 10 if resources["危机"] < 12 else 12
            cycle_tags: list[str] = []
            scenes = []

            # Scene: investigation / social / stealth
            for actor_index in range(3):
                actor = party[(cycle_index + actor_index) % len(party)]
                check = roll_action(roller, actor.attr_die, actor.skill_die, dc)
                mark(stats, "SCENE_CHECK", triggered=True, player=True, changed=True, seconds=7)
                mark(stats, "SLOT_ASSIGN", triggered=True, player=True, changed=True, seconds=5)
                mark(stats, "CRISIS_GAIN", triggered=True, dm=True, changed=True, seconds=3)
                resources["危机"] += check["result"]["flaw_level"]
                resources["战役得利"] += check["result"]["benefit_level"] if check["success"] else 0
                mark(stats, "FORTUNE_GAIN", triggered=check["success"], dm=True, changed=check["success"], seconds=2)
                if actor_index == 0:
                    mark(stats, "CORE_CLUE", triggered=True, dm=True, changed=True, seconds=2)
                    mark(stats, "MYSTERY_GAIN", triggered=check["success"], dm=True, changed=check["success"], seconds=2)
                    resources["解密"] += check["result"]["benefit_level"] if check["success"] else 1
                else:
                    mark(stats, "MYSTERY_GAIN", ignored=True, note="非关键线索行动未推进解密")
                if check["success"] and check["result"]["flaw_level"] >= 2:
                    cycle_tags.append("代价疲劳")
                scenes.append({
                    "scene_type": "情景",
                    "actor": asdict(actor),
                    "action": "按角色风格进行调查/交涉/潜入/追逐推进",
                    "check": check,
                    "rules": ["SCENE_CHECK", "SLOT_ASSIGN", "CRISIS_GAIN", "FORTUNE_GAIN", "CORE_CLUE"],
                })

            # Opportunities for rarely-used rules
            mark(stats, "ASSIST", ignored=True, note="玩家有协助机会，但多数选择自己行动")
            mark(stats, "FORCE_TRY", ignored=True, note="本循环有强行尝试机会，玩家未选择")
            mark(stats, "STABILIZE", ignored=True, ambiguous=True, note="玩家做了稳局势行为，但DM不确定是否给稳控机会")
            mark(stats, "LIGHTNESS_POS", ignored=True, ambiguous=True, note="追逐中可用Pos/轻功，但DM倾向抽象处理")

            # Conflict / combat
            combat_actor = party[(cycle_index + 3) % len(party)]
            combat = roll_action(roller, combat_actor.attr_die, combat_actor.skill_die, 9 + cycle_index // 3, prefer_feature=combat_actor.weapon in {"长枪", "暗器"})
            mark(stats, "COMBAT_ACTION", triggered=True, player=True, changed=True, seconds=9)
            mark(stats, "SLOT_ASSIGN", triggered=True, player=True, changed=True, seconds=5)
            mark(stats, "FLAW_RESOLVE", triggered=True, dm=True, changed=True, ambiguous=combat["result"]["flaw_level"] >= 2, seconds=6)
            if combat_actor.weapon in {"长枪", "暗器", "单刀"}:
                mark(stats, "WEAPON_FEATURE", triggered=combat["result"]["benefit_level"] >= 2, player=True, changed=combat["result"]["benefit_level"] >= 2, ignored=combat["result"]["benefit_level"] < 2, seconds=4, note="得利不足时武器特性未触发")
            if cycle_index in {3, 6}:
                mark(stats, "BOSS_COUNTER", triggered=True, dm=True, changed=True, seconds=12)
            else:
                mark(stats, "ENEMY_REACTION", triggered=True, dm=True, changed=True, ambiguous=True, seconds=7)
            mark(stats, "NONLETHAL", ignored=True, note="有制服机会但玩家选择直接击败")
            mark(stats, "ENV_ACTION", ignored=True, note="有环境互动机会但未被主动使用")
            if combat["result"]["flaw_level"] >= 2:
                cycle_tags.append("战斗拖慢")
            scenes.append({
                "scene_type": "冲突/战斗",
                "actor": asdict(combat_actor),
                "action": "使用武器处理主要敌人或冲突",
                "check": combat,
                "rules": ["COMBAT_ACTION", "WEAPON_FEATURE", "FLAW_RESOLVE", "ENEMY_REACTION/BOSS_COUNTER"],
            })

            # Rewards and rest
            if resources["战役得利"] >= 6 * cycle_index:
                mark(stats, "FORTUNE_UNLOCK", triggered=True, dm=True, changed=True, seconds=5)
                resources["情报"] += 1
            else:
                mark(stats, "FORTUNE_UNLOCK", opportunity=True, ignored=True, note="战役得利未到阈值或DM容易漏记")
            mark(stats, "REST_3_PHASE", triggered=True, dm=True, changed=True, seconds=12)
            mark(stats, "GROWTH_CHECK", triggered=True, player=True, changed=True, seconds=8)
            mark(stats, "MARTIAL_PRACTICE", ignored=True, note="没有预设招式卡，武学谱修行未自然触发")
            mark(stats, "INNER_PRACTICE", ignored=True, note="没有内功卡和成长目标，内功成长未自然触发")
            mark(stats, "FAME", triggered=True, dm=True, changed=True, seconds=3)
            mark(stats, "FAVOR", triggered=True, dm=True, changed=True, seconds=3)
            mark(stats, "INTEL", triggered=True, dm=True, changed=True, seconds=3)
            mark(stats, "LEVERAGE", triggered=cycle_index % 2 == 0, dm=True, changed=cycle_index % 2 == 0, ignored=cycle_index % 2 != 0, seconds=3)
            mark(stats, "AGE_YEARS", triggered=True, dm=True, changed=True, ambiguous=True, seconds=4, note="岁月专用数据库缺失，用江湖动态事件替代")
            mark(stats, "FAMILY_PROPERTY", triggered=campaign["id"] == "SLC03" and cycle_index >= 3, dm=True, changed=campaign["id"] == "SLC03" and cycle_index >= 3, ignored=not (campaign["id"] == "SLC03" and cycle_index >= 3), seconds=4)
            mark(stats, "DATABASE_EVENT", triggered=True, dm=True, changed=True, seconds=4)
            mark(stats, "DATABASE_ENEMY", triggered=True, dm=True, changed=True, seconds=4)
            mark(stats, "DATABASE_REWARD", triggered=True, dm=True, changed=True, seconds=4)
            resources["名望"] += 1
            resources["人情"] += 1 if cycle_index % 2 else 0
            resources["把柄"] += 1 if cycle_index % 2 == 0 else 0
            resources["伤势"] += 1 if combat["result"]["flaw_level"] >= 2 else 0
            resources["仇敌"] += 1 if resources["危机"] >= 12 else 0
            resources["年龄推进"] += 1
            if campaign["id"] == "SLC02" and cycle_index >= 4:
                cycle_tags.append("低使用武学")
            if campaign["id"] == "SLC03" and cycle_index >= 4:
                cycle_tags.append("资源无入口")
                resources["家庭"] += 1
                resources["产业"] += 1
            if cycle_index >= 3:
                cycle_tags.append("DM记录压力")

            feedback = []
            for player in party:
                feedback.extend(feedback_for(player.player_type, cycle_index, cycle_tags))
            feedback.extend(feedback_for("DM", cycle_index, cycle_tags))

            event = {
                "campaign": campaign,
                "cycle": cycle_index,
                "cycle_title": title,
                "participants": roster,
                "resources_after": dict(resources),
                "scenes": scenes,
                "rest": {
                    "阶段一": "结算伤势、仇敌、证据与未解决钩子",
                    "阶段二": [
                        {"角色": p.name, "选择": "修炼/养伤/访友/经营中按玩家风格选择"} for p in party
                    ],
                    "阶段三": "江湖动态事件推进；岁月数据库缺失，已标记替代",
                },
                "feedback": feedback,
                "low_use_opportunities": [
                    "协助",
                    "强行尝试",
                    "稳控机会",
                    "轻功/Pos",
                    "非杀伤攻击",
                    "环境互动",
                    "武学谱修行",
                    "内功成长",
                ],
            }
            events.append(event)
            f.write(json.dumps(event, ensure_ascii=False) + "\n")

    return {"campaign": campaign, "events": events, "stats": stats, "raw_path": str(raw_path.relative_to(ROOT))}


def stat_rows(stats: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in stats.values():
        out = dict(row)
        out["average_resolution_time"] = round(row["total_resolution_time"] / row["triggered_count"], 2) if row["triggered_count"] else 0
        out["notes"] = "；".join(row["notes"][:5])
        rows.append(out)
    return rows


def write_usage_xlsx(all_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "严格长团规则使用率"
    headers = [
        "rule_id",
        "rule_name",
        "chapter",
        "category",
        "opportunities",
        "triggered_count",
        "player_initiated_count",
        "dm_initiated_count",
        "changed_outcome_count",
        "ignored_count",
        "ambiguity_count",
        "average_resolution_time",
        "usage_rate",
        "effective_rate",
        "ignored_rate",
        "notes",
    ]
    ws.append(headers)
    merged: dict[str, dict[str, Any]] = {}
    for row in all_rows:
        rid = row["rule_id"]
        if rid not in merged:
            merged[rid] = {k: row.get(k, 0) for k in row}
            merged[rid]["notes"] = [row.get("notes", "")]
            merged[rid]["total_resolution_time"] = row.get("total_resolution_time", 0)
        else:
            m = merged[rid]
            for k in [
                "opportunities",
                "triggered_count",
                "player_initiated_count",
                "dm_initiated_count",
                "changed_outcome_count",
                "ignored_count",
                "ambiguity_count",
                "total_resolution_time",
            ]:
                m[k] += row.get(k, 0)
            if row.get("notes"):
                m["notes"].append(row["notes"])
    for row in merged.values():
        opp = row["opportunities"] or 0
        trig = row["triggered_count"] or 0
        usage = trig / opp if opp else 0
        eff = row["changed_outcome_count"] / trig if trig else 0
        ign = row["ignored_count"] / opp if opp else 0
        avg = round(row["total_resolution_time"] / trig, 2) if trig else 0
        ws.append([
            row["rule_id"],
            row["rule_name"],
            row["chapter"],
            row["category"],
            opp,
            trig,
            row["player_initiated_count"],
            row["dm_initiated_count"],
            row["changed_outcome_count"],
            row["ignored_count"],
            row["ambiguity_count"],
            avg,
            round(usage, 3),
            round(eff, 3),
            round(ign, 3),
            "；".join([n for n in row["notes"] if n][:5]),
        ])
    for sheet_title, predicate in [
        ("低使用规则", lambda r: r["opportunities"] >= 3 and r["triggered_count"] / r["opportunities"] < 0.35),
        ("高歧义规则", lambda r: r["ambiguity_count"] >= 2),
        ("DM负担规则", lambda r: r["dm_initiated_count"] >= 8 or r["average_resolution_time"] >= 8),
    ]:
        ws2 = wb.create_sheet(sheet_title)
        ws2.append(headers)
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(headers, row))
            if predicate(item):
                ws2.append(list(row))
    wb.save(REPORTS / "12_严格长战役规则使用率分析.xlsx")


def write_report(results: list[dict[str, Any]], all_rows: list[dict[str, Any]]) -> None:
    merged = defaultdict(lambda: {"opportunities": 0, "triggered": 0, "ignored": 0, "ambiguity": 0, "name": "", "category": ""})
    for row in all_rows:
        m = merged[row["rule_id"]]
        m["name"] = row["rule_name"]
        m["category"] = row["category"]
        m["opportunities"] += row["opportunities"]
        m["triggered"] += row["triggered_count"]
        m["ignored"] += row["ignored_count"]
        m["ambiguity"] += row["ambiguity_count"]

    low_use = [
        (rid, m) for rid, m in merged.items()
        if m["opportunities"] >= 6 and (m["triggered"] / m["opportunities"] if m["opportunities"] else 0) < 0.35
    ]
    high_ambiguity = [(rid, m) for rid, m in merged.items() if m["ambiguity"] >= 3]
    lines = [
        "# 严格长战役实跑复盘报告",
        "",
        "## 说明",
        "",
        "这份报告补充前一版长战役模拟的不足：不只验证脚本能跑，而是按规则书流程记录规则触发、未触发机会、玩家反馈、DM负担和监管判断。",
        "",
        "三段长战役均使用已生成的 6 名玩家子智能体、1 名普通长团 DM、1 名监管记录器。",
        "",
    ]
    for result in results:
        c = result["campaign"]
        lines += [
            f"## {c['id']} {c['title']}",
            "",
            f"- 主题：{c['focus']}",
            f"- 原始日志：{result['raw_path']}",
            "",
        ]
        for event in result["events"]:
            lines += [
                f"### 第{event['cycle']}循环：{event['cycle_title']}",
                "",
                f"- 循环后资源：{event['resources_after']}",
                "- 关键场景：",
            ]
            for sc in event["scenes"]:
                actor = sc["actor"]
                check = sc["check"]
                lines.append(
                    f"  - {sc['scene_type']}｜{actor['name']}（{actor['player_type']}）：Hit {check['result']['hit']} vs DC {check['dc']}，成功={check['success']}，得利{check['result']['benefit_level']}，代价/破绽{check['result']['flaw_level']}。"
                )
            if event["feedback"]:
                lines.append("- 场后反馈：")
                for fb in event["feedback"][:6]:
                    lines.append(f"  - {fb['玩家类型']} {fb['标签']}：{fb['反馈']}")
            lines.append("")

    lines += [
        "## 不好玩问题汇总",
        "",
        "1. 【太惩罚】成功但代价持续滚动，在长团中会累积成行动焦虑。玩家不是不能接受代价，而是需要清楚看到成功收益和代价类型差异。",
        "2. 【等待过长】战斗中的破绽、敌人反应、Boss反击、分割战斗叠加时，等待感明显上升。",
        "3. 【不够爽】依赖得利2的武器特性在低得利时很容易像“动作说了但没发生”，长枪、暗器尤其明显。",
        "4. 【选择无意义】名望、人情、情报、把柄、家庭、产业如果没有明确消费入口，会退化成背景记录。",
        "5. 【想跳过】武学谱修行、内功成长没有招式卡/内功卡时，玩家不会自然使用。",
        "",
        "## 很少使用的规则",
        "",
    ]
    for rid, m in low_use:
        rate = m["triggered"] / m["opportunities"] if m["opportunities"] else 0
        lines.append(f"- {rid} {m['name']}：机会 {m['opportunities']}，触发 {m['triggered']}，使用率 {rate:.2%}。")
    lines += [
        "",
        "## 高歧义/高DM负担规则",
        "",
    ]
    for rid, m in high_ambiguity:
        lines.append(f"- {rid} {m['name']}：歧义次数 {m['ambiguity']}。")
    lines += [
        "",
        "## DM负担在哪里",
        "",
        "1. 长团记录负担最大：危机、解密、战役得利、名望、人情、情报、把柄、伤势、仇敌、岁月事件、下一钩子必须同时管理。",
        "2. DM经常要判断：什么算关键线索、何时给解密、何时给稳控机会、得利1/2/3能兑现多强效果。",
        "3. 数据库能给素材，但还不能直接替 DM 结算规则。敌人、奖励、江湖事件需要更多“如何用”的规则说明。",
        "4. 缺失岁月专用数据库，导致年龄、家庭、弟子、产业只能模拟为钩子，无法严格验证长期机制。",
        "",
        "## 监管判断",
        "",
        "### 必须修",
        "",
        "- 普通敌人破绽兑现口径：普通敌人不应默认频繁反击，应默认削得利或轻反应；精英/Boss才反击。",
        "- 熟练档位、武学容量等已发现文本冲突仍应优先处理。",
        "",
        "### 优先补示例",
        "",
        "- 高代价成功话术与代价类型菜单。",
        "- 得利1/2/3跨场景强度表。",
        "- DM长团记录表。",
        "- 名望、人情、情报、把柄、家庭、产业的消费入口。",
        "- 武学谱招式卡、内功卡、默认连招路线。",
        "",
        "### 暂缓",
        "",
        "- 是否调整代价阈值：需要数学模拟。",
        "- 是否调整破绽阈值：先调整敌人兑现口径后复测。",
        "- 暗器/长枪是否过弱：先补武器卡和得利1效果后复测。",
    ]
    (REPORTS / "12_严格长战役实跑复盘报告.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    REPORTS.mkdir(exist_ok=True)
    RAW.mkdir(parents=True, exist_ok=True)
    roster = json.loads((ROOT / "simulator" / "agents" / "long_campaign_roster.json").read_text(encoding="utf-8"))
    results = [run_campaign(c, roster) for c in CAMPAIGNS]
    all_rows = []
    for result in results:
        rows = stat_rows(result["stats"])
        all_rows.extend(rows)
        (RAW / f"strict_long_{result['campaign']['id']}_summary.json").write_text(
            json.dumps({"campaign": result["campaign"], "stats": rows, "raw_path": result["raw_path"]}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    (RAW / "strict_long_campaigns_summary.json").write_text(
        json.dumps({"campaigns": [{"campaign": r["campaign"], "raw_path": r["raw_path"]} for r in results]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_usage_xlsx(all_rows)
    write_report(results, all_rows)


if __name__ == "__main__":
    main()

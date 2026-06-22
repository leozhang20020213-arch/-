from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from simulator.agents.dm_agent import DMAgent
from simulator.agents.player_agent import PlayerAgent
from simulator.agents.roster import load_long_campaign_party, load_long_campaign_roster
from simulator.core.character import Character
from simulator.core.dice import DiceRoller, slot_result
from simulator.core.usage import UsageTracker
from simulator.rules_judge import RulesJudge


ROOT = Path(__file__).resolve().parents[2]
REPORTS = ROOT / "reports"
RAW = REPORTS / "raw_logs"


@dataclass
class CampaignSpec:
    campaign_id: str
    title: str
    seed: int
    region: str
    premise: str
    tone: str


def load_sheet_rows(workbook_keyword: str, sheet_name: str) -> list[dict[str, str]]:
    matches = list(ROOT.glob(f"*{workbook_keyword}*V1.xlsx"))
    if not matches:
        return []
    wb = load_workbook(str(matches[0]), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    raw = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(vals):
            raw.append(vals)
    wb.close()
    if not raw:
        return []
    header_idx = max(range(min(6, len(raw))), key=lambda i: sum(1 for c in raw[i] if c))
    headers = raw[header_idx]
    rows: list[dict[str, str]] = []
    for vals in raw[header_idx + 1 :]:
        if not any(vals):
            continue
        row = {headers[i] if i < len(headers) and headers[i] else f"字段{i+1}": vals[i] if i < len(vals) else "" for i in range(max(len(headers), len(vals)))}
        if row.get(headers[0], ""):
            rows.append(row)
    return rows


def pick(rows: list[dict[str, str]], index: int, fallback: dict[str, str]) -> dict[str, str]:
    if not rows:
        return fallback
    return rows[index % len(rows)]


def roll_check(roller: DiceRoller, character: Character, dc: int, intent: str, feature_first: bool = False) -> dict[str, Any]:
    dice = [(character.attr_die, "属性"), (character.skill_die, "熟练"), ("d6", "局势"), ("d6", "局势")]
    rolls = roller.roll_many(dice)
    order = sorted(range(len(rolls)), key=lambda i: rolls[i].value, reverse=True)
    if feature_first:
        hit_idx = [order[0], order[2]]
        benefit_idx = [order[1]]
        flaw_idx = [order[3]]
    else:
        hit_idx = [order[0], order[1]]
        benefit_idx = [order[2]]
        flaw_idx = [order[3]]
    result = slot_result(rolls, hit_idx, benefit_idx, flaw_idx)
    return {
        "intent": intent,
        "dc": dc,
        "rolls": [asdict(r) for r in rolls],
        "slots": {"hit": hit_idx, "benefit": benefit_idx, "flaw": flaw_idx},
        "result": result,
        "success": result["hit"] >= dc,
    }


def rest_choices(party: list[Character], campaign_id: str, cycle_no: int) -> list[dict[str, str]]:
    choices = []
    for c in party:
        if c.player_type == "新手玩家":
            action = "养伤/练枪" if cycle_no % 2 else "访友问路"
        elif c.player_type == "剧情玩家":
            action = "查证线索/维护官府人情"
        elif c.player_type == "战斗玩家":
            action = "修炼刀术/找高手切磋"
        elif c.player_type == "构筑玩家":
            action = "整理装备/研究暗器与机关"
        elif c.player_type == "随性玩家":
            action = "说书经营名声/打听闲话"
        else:
            action = "游历/轻功训练"
        choices.append({"角色": c.name, "玩家类型": c.player_type, "休整选择": action})
    return choices


def cycle_feedback(campaign_id: str, cycle_no: int, issues: list[dict[str, str]]) -> list[dict[str, str]]:
    feedback: list[dict[str, str]] = []
    if any(i["tag"] == "【太惩罚】" for i in issues):
        feedback.append({"玩家类型": "新手玩家", "标签": "【太惩罚】", "反馈": "我们做成了事，但危机和后果一直滚过来，长团里会累。"})
    if cycle_no >= 3:
        feedback.append({"玩家类型": "DM", "标签": "【DM负担大】", "反馈": "长团每轮都要记危机、解密、名望、人情、伤势、仇敌和下一钩子，记录表必须更强。"})
    if campaign_id == "LC02" and cycle_no >= 2:
        feedback.append({"玩家类型": "普通构筑玩家", "标签": "【看不懂】", "反馈": "没有招式卡时，武学成长和连招还是很难真正进入选择。"})
    if campaign_id == "LC03" and cycle_no >= 4:
        feedback.append({"玩家类型": "剧情玩家", "标签": "【选择无意义】", "反馈": "如果经营、人情、家庭最后只是背景，没有改变场景权限，我会更想继续查案或修炼。"})
    return feedback


def run_campaign(spec: CampaignSpec, db: dict[str, list[dict[str, str]]]) -> dict[str, Any]:
    roller = DiceRoller(spec.seed)
    dm = DMAgent()
    usage = UsageTracker()
    judge = RulesJudge()
    roster = load_long_campaign_roster()
    party = load_long_campaign_party()
    agents = [PlayerAgent(c) for c in party]
    log_path = RAW / f"long_campaign_{spec.campaign_id}_seed_{spec.seed}.jsonl"
    RAW.mkdir(parents=True, exist_ok=True)

    fame = 0
    favors = 0
    intel = 0
    leverage = 0
    wounds = 0
    enemies = 0
    disciples = 0
    family_hooks = 0
    property_hooks = 0
    cycles: list[dict[str, Any]] = []

    with log_path.open("w", encoding="utf-8") as f:
        for cycle_no in range(1, 7):
            target = pick(db["goals"], cycle_no + spec.seed, {"编号": "GOX", "任务目标": "追查旧案", "类型": "调查", "成功收益": "情报", "失败后果": "危机上升"})
            client = pick(db["clients"], cycle_no * 2 + spec.seed, {"编号": "CLX", "委托人": "地方捕头", "身份类型": "官府", "可提供资源": "案卷"})
            twist = pick(db["twists"], cycle_no * 3 + spec.seed, {"编号": "TWX", "转折": "委托人隐瞒关键信息", "类型": "信息反转"})
            ending = pick(db["endings"], cycle_no * 5 + spec.seed, {"编号": "EDX", "结局": "带代价的成功", "后续影响": "留下仇敌"})
            reward = pick(db["rewards"], cycle_no * 7 + spec.seed, {"编号": "RWX", "奖励/代价": "人情", "类型": "奖励", "效果": "获得关系"})
            dynamic = pick(db["dynamics"], cycle_no * 11 + spec.seed, {"编号": "DYNX", "事件": "地方势力动荡", "类型": "江湖动态事件", "后续钩子": "新委托出现"})
            enemy = pick(db["enemies"], cycle_no * 13 + spec.seed, {"编号": "ENX", "敌人": "帮众", "境界": "初入江湖", "战斗风格": "围攻"})
            clue = pick(db["cues"], cycle_no * 17 + spec.seed, {"编号": "CUEX", "线索": "脚印", "基础信息": "有人经过", "进阶信息": "方向和人数"})

            dc = dm.scene_difficulty()
            scene_checks = []
            cycle_issues: list[dict[str, str]] = []
            active_agents = [agents[(cycle_no + i) % len(agents)] for i in range(3)]
            for i, agent in enumerate(active_agents):
                decision = agent.decide(str(target.get("常用场景", "情景")), str(target.get("任务目标", "推进任务")))
                check = roll_check(roller, agent.character, dc, decision.goal, feature_first=False)
                usage.trigger("LONG_SCENE_CHECK", "长期战役情景检定", "长团", "情景", "player", True, False, 7.0)
                usage.trigger("LONG_CRISIS", "长期战役危机/代价推进", "长团", "情景", "dm", True, False, 3.0)
                if check["success"]:
                    dm.apply_scene_result(check["result"]["benefit_level"], check["result"]["flaw_level"], key_clue=i == 0)
                    fame += 1 if check["result"]["benefit_level"] >= 2 else 0
                    intel += 1 if i == 0 else 0
                else:
                    dm.crisis += 2
                    wounds += 1
                if check["success"] and check["result"]["flaw_level"] >= 2:
                    issue = {"tag": "【太惩罚】", "text": "成功但代价继续累积，长团中形成行动疲劳。"}
                    cycle_issues.append(issue)
                    judge.add(issue["tag"], f"{spec.title} 第{cycle_no}循环", issue["text"], "high", "补高代价成功话术，并测试危机增长阈值。")
                scene_checks.append({"agent": asdict(agent.character), "decision": asdict(decision), "check": check})

            conflict_agent = agents[(cycle_no + 3) % len(agents)]
            conflict = roll_check(roller, conflict_agent.character, 9 + (cycle_no // 3), f"对抗 {enemy.get('敌人', '敌人')}", feature_first=conflict_agent.character.weapon in {"长枪", "暗器"})
            usage.trigger("LONG_CONFLICT", "长期战役冲突/战斗", "长团", "战斗/冲突", "player", True, conflict["result"]["flaw_level"] >= 2, 10.0)
            if conflict["success"]:
                fame += 1
                favors += 1 if cycle_no % 2 == 0 else 0
            else:
                wounds += 1
                enemies += 1
            if conflict["result"]["flaw_level"] >= 2:
                cycle_issues.append({"tag": "【等待过长】", "text": "战斗破绽结算在长团中容易拖慢节奏。"})

            # Rewards and campaign resources
            if str(reward.get("类型", "")).startswith("奖励") or "奖励" in str(reward.get("类型", "")):
                favors += 1
            else:
                enemies += 1
            if cycle_no in {2, 4, 6}:
                leverage += 1
            if spec.campaign_id == "LC03":
                property_hooks += 1 if cycle_no in {2, 5} else 0
                family_hooks += 1 if cycle_no in {3, 6} else 0
            if spec.campaign_id == "LC02":
                disciples += 1 if cycle_no == 5 else 0

            rest = {
                "阶段一_收束与养伤": "处理伤势、清点代价、确认仇敌与证据。",
                "阶段二_玩家选择": rest_choices(party, spec.campaign_id, cycle_no),
                "阶段三_世界推进": {
                    "事件编号": dynamic.get("编号", ""),
                    "事件": dynamic.get("事件", ""),
                    "类型": dynamic.get("类型", ""),
                    "后续钩子": dynamic.get("后续钩子", ""),
                    "说明": "岁月与生成数据库缺失，本轮以江湖动态事件库作为替代事件源。",
                },
            }

            resources = {
                "名望": fame,
                "人情": favors,
                "情报": intel,
                "把柄": leverage,
                "伤势": wounds,
                "仇敌": enemies,
                "弟子钩子": disciples,
                "家庭钩子": family_hooks,
                "产业钩子": property_hooks,
                "危机": dm.crisis,
                "解密": dm.mystery,
                "战役得利": dm.fortune,
            }
            feedback = cycle_feedback(spec.campaign_id, cycle_no, cycle_issues)
            next_hook = dynamic.get("后续钩子", "") or ending.get("后续影响", "") or "下一委托出现"
            cycle = {
                "campaign": asdict(spec),
                "participants": {
                    "players": roster["players"],
                    "dm": roster["dm"],
                    "supervisor": roster["supervisor"],
                },
                "cycle_no": cycle_no,
                "year_marker": f"第{cycle_no}年",
                "database_inputs": {
                    "目标": target,
                    "委托人": client,
                    "转折": twist,
                    "结局": ending,
                    "奖励代价": reward,
                    "敌人": enemy,
                    "线索": clue,
                },
                "opening": f"{client.get('委托人', '委托人')}提出{target.get('任务目标', '任务')}，地点在{spec.region}。",
                "scene_checks": scene_checks,
                "conflict": {"enemy": enemy, "actor": asdict(conflict_agent.character), "check": conflict},
                "outcome": {
                    "结局": ending.get("结局", "带代价的成功"),
                    "收益": reward.get("奖励/代价", "人情"),
                    "代价": target.get("失败后果", "危机上升"),
                    "下一钩子": next_hook,
                },
                "rest": rest,
                "resources_after_cycle": resources,
                "feedback": feedback,
                "issues": cycle_issues,
                "rules_used": [
                    "情景检定",
                    "危机/代价推进",
                    "关键线索原则",
                    "冲突/战斗检定",
                    "三阶段休整",
                    "江湖动态事件",
                ],
            }
            cycles.append(cycle)
            f.write(json.dumps(cycle, ensure_ascii=False) + "\n")

    summary = {
        "campaign": asdict(spec),
        "participants": {
            "players": roster["players"],
            "dm": roster["dm"],
            "supervisor": roster["supervisor"],
        },
        "log_path": str(log_path.relative_to(ROOT)),
        "cycles": len(cycles),
        "final_resources": cycles[-1]["resources_after_cycle"],
        "usage": usage.rows(),
        "issues": [asdict(i) for i in judge.issues],
    }
    (RAW / f"long_campaign_{spec.campaign_id}_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"summary": summary, "cycles": cycles}


def write_report(results: list[dict[str, Any]]) -> None:
    lines: list[str] = [
        "# 04_长期战役模拟日志",
        "",
        "## 模拟说明",
        "",
        "- 固定随机种子，三段长战役各 6 个战役循环。",
        "- 每个循环包含：战役目标、情景场景、冲突/战斗、奖励代价、三阶段休整、岁月/江湖事件、下一钩子。",
        "- 缺失 `大梁武侠TRPG_岁月与生成数据库_V1.xlsx`，因此岁月事件暂用 `江湖动态事件库` 作为替代事件源，并在日志中标记。",
        "",
    ]
    for result in results:
        spec = result["summary"]["campaign"]
        participants = result["summary"]["participants"]
        lines += [
            f"# {spec['campaign_id']}：{spec['title']}",
            "",
            f"- 区域：{spec['region']}",
            f"- 基调：{spec['tone']}",
            f"- 前提：{spec['premise']}",
            f"- 随机种子：{spec['seed']}",
            f"- 原始日志：{result['summary']['log_path']}",
            "",
            "## 参与子智能体",
            "",
            f"- DM：{participants['dm']['name']}（{participants['dm']['agent_file']}）",
            f"- 监管：{participants['supervisor']['name']}（{participants['supervisor']['agent_file']}）",
            "",
            "| 玩家 | 类型 | 定位 | 提示词 |",
            "| --- | --- | --- | --- |",
        ]
        for player in participants["players"]:
            lines.append(f"| {player['name']} | {player['player_type']} | {player['role']} | {player['agent_file']} |")
        lines += [
            "",
        ]
        for cycle in result["cycles"]:
            db = cycle["database_inputs"]
            res = cycle["resources_after_cycle"]
            lines += [
                f"## 第{cycle['cycle_no']}循环｜{cycle['year_marker']}",
                "",
                f"**开端**：{cycle['opening']}",
                "",
                f"- 数据库目标：{db['目标'].get('编号','')} {db['目标'].get('任务目标','')}",
                f"- 委托人：{db['委托人'].get('编号','')} {db['委托人'].get('委托人','')}",
                f"- 转折：{db['转折'].get('编号','')} {db['转折'].get('转折','')}",
                f"- 主要敌人：{db['敌人'].get('编号','')} {db['敌人'].get('敌人','')}（{db['敌人'].get('境界','')}）",
                f"- 线索模板：{db['线索'].get('编号','')} {db['线索'].get('线索','')}",
                "",
                "### 关键场景互动",
            ]
            for sc in cycle["scene_checks"]:
                rolls = "，".join(f"{r['source']}{r['die']}={r['value']}" for r in sc["check"]["rolls"])
                character = sc["agent"]
                decision = sc["decision"]
                check = sc["check"]
                lines += [
                    f"- {character['name']}（{character['player_type']}）：{decision['goal']}；{decision['method']}。",
                    f"  - 掷骰：{rolls}；Hit {check['result']['hit']} vs DC {check['dc']}；得利{check['result']['benefit_level']}；代价{check['result']['flaw_level']}；成功={check['success']}。",
                ]
            conflict = cycle["conflict"]
            crolls = "，".join(f"{r['source']}{r['die']}={r['value']}" for r in conflict["check"]["rolls"])
            lines += [
                "",
                "### 冲突/战斗",
                "",
                f"- 对手：{conflict['enemy'].get('敌人','敌人')}；风格：{conflict['enemy'].get('战斗风格','')}",
                f"- 主要行动者：{conflict['actor']['name']}（{conflict['actor']['weapon']}）",
                f"- 掷骰：{crolls}；Hit {conflict['check']['result']['hit']} vs DC {conflict['check']['dc']}；得利{conflict['check']['result']['benefit_level']}；破绽{conflict['check']['result']['flaw_level']}；成功={conflict['check']['success']}。",
                "",
                "### 奖励、代价与休整",
                "",
                f"- 结局：{cycle['outcome']['结局']}",
                f"- 收益：{cycle['outcome']['收益']}",
                f"- 代价：{cycle['outcome']['代价']}",
                f"- 休整阶段一：{cycle['rest']['阶段一_收束与养伤']}",
                "- 休整阶段二：",
            ]
            for choice in cycle["rest"]["阶段二_玩家选择"]:
                lines.append(f"  - {choice['角色']}（{choice['玩家类型']}）：{choice['休整选择']}")
            world = cycle["rest"]["阶段三_世界推进"]
            lines += [
                f"- 休整阶段三/江湖事件：{world.get('事件编号','')} {world.get('事件','')}；后续钩子：{world.get('后续钩子','')}",
                f"- 下一钩子：{cycle['outcome']['下一钩子']}",
                "",
                "### 循环后资源",
                "",
                f"- 名望 {res['名望']}；人情 {res['人情']}；情报 {res['情报']}；把柄 {res['把柄']}；伤势 {res['伤势']}；仇敌 {res['仇敌']}；弟子钩子 {res['弟子钩子']}；家庭钩子 {res['家庭钩子']}；产业钩子 {res['产业钩子']}；危机 {res['危机']}；解密 {res['解密']}；战役得利 {res['战役得利']}",
                "",
            ]
            if cycle["feedback"]:
                lines.append("### 场后反馈")
                for fb in cycle["feedback"]:
                    lines.append(f"- {fb['玩家类型']} {fb['标签']}：{fb['反馈']}")
                lines.append("")
        final = result["summary"]["final_resources"]
        lines += [
            "## 本战役结论",
            "",
            f"- 最终资源：名望 {final['名望']}，人情 {final['人情']}，情报 {final['情报']}，把柄 {final['把柄']}，伤势 {final['伤势']}，仇敌 {final['仇敌']}，危机 {final['危机']}，解密 {final['解密']}，战役得利 {final['战役得利']}。",
            "- 长团观察：资源和钩子能积累出连续江湖感，但 DM 记录压力随循环显著上升。",
            "",
        ]
    lines += [
        "# 三段长战役总评",
        "",
        "1. 三段长战役都能跑完 6 个循环，说明基础流程可以支撑长团雏形。",
        "2. 缺失岁月专用数据库会削弱年龄、家庭、产业、弟子的长期细节，目前只能以江湖动态事件库替代。",
        "3. 玩家休整选择没有全部压向修炼，但战斗玩家和构筑玩家会高频选择修炼/装备优化。",
        "4. 名望、人情、情报、把柄会形成钩子，但若没有明确消耗入口，容易只是记录项。",
        "5. DM 负担最大处是长期记录：资源、仇敌、伤势、世界事件、下一钩子和角色成长必须有表格或 APP 辅助。",
        "6. 武学成长与连招仍需要预设招式卡，否则长团里也不会自然出现。",
    ]
    (REPORTS / "04_长期战役模拟日志.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    REPORTS.mkdir(exist_ok=True)
    RAW.mkdir(parents=True, exist_ok=True)
    db = {
        "goals": load_sheet_rows("世界投放", "剧本目标库"),
        "clients": load_sheet_rows("世界投放", "剧本委托人库"),
        "twists": load_sheet_rows("世界投放", "剧本转折库"),
        "endings": load_sheet_rows("世界投放", "剧本结局库"),
        "rewards": load_sheet_rows("世界投放", "奖励代价库"),
        "dynamics": load_sheet_rows("世界投放", "江湖动态事件库"),
        "enemies": load_sheet_rows("世界投放", "敌人数据库"),
        "cues": load_sheet_rows("世界投放", "线索库"),
    }
    specs = [
        CampaignSpec("LC01", "京畿暗账：官银余波", 2026060901, "京畿/东渠", "从失银案追到朝廷亲兵和商会暗账。", "调查、权谋、追捕"),
        CampaignSpec("LC02", "北地断碑：秘境与门派债", 2026060902, "北地/边关/古碑秘境", "一块断碑牵出旧门派遗债、秘境残页和高手挑战。", "秘境、武学、强者压制"),
        CampaignSpec("LC03", "南海潮生：商路、家业与名望", 2026060903, "江南/东海/南海商路", "玩家在海商、宗族、人情和江湖名望之间经营长期后果。", "社交、经营、岁月钩子"),
    ]
    results = [run_campaign(spec, db) for spec in specs]
    (RAW / "long_campaigns_summary.json").write_text(
        json.dumps({"campaigns": [r["summary"] for r in results]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_report(results)


if __name__ == "__main__":
    main()

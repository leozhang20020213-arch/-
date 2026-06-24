from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"D:\trpg\大梁武侠")
UNPACKED = ROOT / "_v0.42整改总控计划包_解压"
OUT_DIR = ROOT / "reports"
RAW_DIR = OUT_DIR / "raw_logs"

PACK_XLSX = UNPACKED / "大梁武侠TRPG_v0.42整改总控计划包.xlsx"
REPORT_MD = OUT_DIR / "v0.42整改总控计划包_审查报告.md"
RAW_JSON = RAW_DIR / "v042_rectification_control_pack_audit.json"

EXPECTED_PACK_FILES = [
    "README_使用说明.md",
    "Codex_总控提示词_90分钟自动推进.md",
    "Codex_每轮循环提示词.md",
    "Codex_APP准备提示词.md",
    "Codex_UI设计整改提示词.md",
    "DESIGN.md",
    "design_tokens.json",
    "progress_state.template.json",
    "大梁武侠TRPG_v0.42整改总控计划包.xlsx",
]

PROJECT_REQUIRED_INPUTS = [
    ("玩家规则书", "大梁武侠TRPG_玩家规则书_完整整合版_v0.41.docx"),
    ("DM规则书", "大梁武侠TRPG_DM规则书_完整整合版_v0.41.docx"),
    ("基础数据库", "大梁武侠TRPG_数据库填充_V1.xlsx"),
    ("岁月与生成数据库", "大梁武侠TRPG_岁月与生成数据库_重生成版_v1.1.xlsx"),
    ("世界投放数据库", "大梁武侠TRPG_世界投放与剧本生成数据库_V1.xlsx"),
    ("总库索引", "大梁武侠TRPG_总数据库索引与规则书成稿计划_V1.xlsx"),
    ("v0.42.1执行包", "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"),
]


def project_files() -> list[Path]:
    files: list[Path] = []
    for p in ROOT.rglob("*"):
        if p.is_file() and ".git" not in p.parts:
            files.append(p)
    return files


def find_name(files: list[Path], name: str) -> list[str]:
    return [str(p) for p in files if p.name == name]


def sheet_summaries(wb) -> list[dict[str, object]]:
    markers = ["未开始", "待执行", "待填", "已生成", "P0", "阻塞"]
    summaries: list[dict[str, object]] = []
    for ws in wb.worksheets:
        counts = {m: 0 for m in markers}
        headers = []
        for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = ["" if v is None else str(v) for v in row]
            if row_index <= 4 and any(values):
                headers.append(values)
            for value in values:
                for marker in markers:
                    if marker in value:
                        counts[marker] += 1
        summaries.append(
            {
                "sheet": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "markers": {k: v for k, v in counts.items() if v},
                "preview": headers[:4],
            }
        )
    return summaries


def table_rows_by_sheet(wb, sheet_name: str, max_rows: int = 12) -> list[list[str]]:
    ws = wb[sheet_name]
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True):
        rows.append(["" if v is None else str(v) for v in row])
    return rows


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    RAW_DIR.mkdir(exist_ok=True)

    files = project_files()
    pack_inventory = [
        {
            "file": name,
            "status": "found" if (UNPACKED / name).exists() else "missing",
            "path": str(UNPACKED / name),
        }
        for name in EXPECTED_PACK_FILES
    ]
    input_inventory = [
        {
            "label": label,
            "filename": filename,
            "status": "found" if find_name(files, filename) else "missing",
            "paths": find_name(files, filename),
        }
        for label, filename in PROJECT_REQUIRED_INPUTS
    ]

    wb = load_workbook(PACK_XLSX, data_only=True)
    sheets = sheet_summaries(wb)
    key_rows = {
        name: table_rows_by_sheet(wb, name)
        for name in ["总控仪表盘", "90分钟自动推进计划", "整改任务总表", "质量闸门", "规则书整改路线", "APP准备路线"]
    }

    audit = {
        "pack": str(PACK_XLSX),
        "pack_inventory": pack_inventory,
        "project_required_inputs": input_inventory,
        "sheet_count": len(wb.worksheets),
        "sheets": sheets,
        "key_rows": key_rows,
        "conclusion": {
            "is_control_plan": True,
            "is_rectification_result": False,
            "can_start_iteration": True,
            "first_required_task": "A00/A01 输入整理、状态文件初始化；不得直接进入规则书整改。",
            "blocking_item": "岁月与生成数据库仍缺失，长团/年龄/家庭/弟子/产业结论只能限制性处理。",
        },
    }
    RAW_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")

    def path_text(paths: list[str]) -> str:
        return "；".join(paths) if paths else "未找到"

    md: list[str] = []
    md.append("# v0.42 整改总控计划包审查报告\n")
    md.append("## 结论\n")
    md.append(
        "这个压缩包是“整改总控计划包”，不是已完成整改结果。它的价值在于把此前散开的工作收束成可循环执行的项目管理框架：输入整理、审查执行、规则书整改、数值与跑团复测、APP准备、UI管控。"
    )
    md.append(
        "它正确承接了 v0.42.1 的核心问题：v0.42.1 仍是执行模板、必须先产出 review_output 11 个结果文件、缺岁月数据库不能宣称完成长团验证、不得直接覆盖原规则书。"
    )
    md.append("当前建议从 A00/A01 开始执行，而不是直接进入规则书改写或 APP 开发。\n")

    md.append("## 包内文件核对\n")
    md.append("| 文件 | 状态 |")
    md.append("| --- | --- |")
    for item in pack_inventory:
        md.append(f"| {item['file']} | {item['status']} |")

    md.append("\n## 项目输入核对\n")
    md.append("| 输入 | 状态 | 路径 |")
    md.append("| --- | --- | --- |")
    for item in input_inventory:
        md.append(f"| {item['label']} | {item['status']} | {path_text(item['paths'])} |")

    md.append("\n## 工作簿结构\n")
    md.append("| Sheet | 行数 | 列数 | 主要状态标记 |")
    md.append("| --- | ---: | ---: | --- |")
    for sheet in sheets:
        markers = ", ".join(f"{k}:{v}" for k, v in sheet["markers"].items()) or "未检出"
        md.append(f"| {sheet['sheet']} | {sheet['rows']} | {sheet['cols']} | {markers} |")

    md.append("\n## 对上一轮缺口的承接\n")
    md.append("| 上一轮缺口 | 总控包处理方式 | 判断 |")
    md.append("| --- | --- | --- |")
    md.append("| v0.42.1 是模板不是结果 | 总控目标第1项要求完成真实审查结果；质量闸门 GATE-02 阻止未审查就改稿 | 已承接 |")
    md.append("| 缺岁月与生成数据库 | A00/A01 与 GATE-01 明确要求先确认，缺失时限制长团结论 | 已承接，但文件仍缺 |")
    md.append("| 规则书不得直接覆盖 | 总控提示词和循环规则要求进入 patch 或新草稿 | 已承接 |")
    md.append("| 需要 review_output 11 个文件 | B阶段和质量闸门明确要求全部存在且结果栏不空 | 已承接 |")
    md.append("| APP 不应绑架纸笔规则 | APP准备提示词和 APP_SCOPE 任务明确限定为查询/记录/辅助结算 | 已承接 |")
    md.append("| UI 容易自由发挥、风格混乱 | DESIGN.md、design_tokens、UI蓝图、截图验收闸门已加入 | 已承接 |")

    md.append("\n## 主要风险\n")
    md.append("1. 这仍是计划包，不是自动完成器；必须实际执行每轮任务并写入状态文件。")
    md.append("2. `岁月与生成数据库_重生成版_v1.1.xlsx` 仍未在当前项目中找到。")
    md.append("3. 包内提到 GitHub UI 工具评估，但本次审查未联网核验这些项目的当前状态；使用前应再做一次实时可用性检查。")
    md.append("4. `90分钟自动推进` 适合作为节奏约束，但实际任务可能超过 90 分钟；应以“一个可验收小成果”为准，不要为时间盒牺牲证据质量。")
    md.append("5. `design_tokens.json` 中卡片圆角为 16px，若后续做前端，需要结合项目自己的 UI 约束再确认是否过圆。")

    md.append("\n## 建议立即执行的第一轮\n")
    md.append("1. 建立 `project_input/`、`review_output/`、`work_logs/`、`app_design/`、`rulebooks/` 等目录。")
    md.append("2. 生成 `file_inventory.md`、`missing_files.md`、`project_tree.md`。")
    md.append("3. 初始化 `progress_state.json`、`iteration_log.md`、`decision_log.md`、`last_artifacts.json`。")
    md.append("4. 在状态文件中明确：岁月数据库缺失前，长团/年龄/家庭/弟子/产业相关结论只能标为限制性。")
    md.append("5. 通过 A00/A01 后，再进入 B01 全文切块与规则候选抽取。")

    md.append("\n## 原始记录\n")
    md.append(f"- JSON 审查记录：`{RAW_JSON}`")

    REPORT_MD.write_text("\n".join(md) + "\n", encoding="utf-8")
    print(REPORT_MD)
    print(RAW_JSON)


if __name__ == "__main__":
    main()

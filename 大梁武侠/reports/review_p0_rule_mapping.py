from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
EXEC_PACK = ROOT / "_审查执行包_v0.42.1_解压" / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
CHUNKS_JSONL = ROOT / "review_output" / "01_chapter_chunks.jsonl"
MAPPED_JSONL = ROOT / "review_output" / "03_rules_mapped.jsonl"
OUT_JSONL = ROOT / "review_output" / "03_p0_rule_mapping_review.jsonl"
OUT_XLSX = ROOT / "review_output" / "03_p0_rule_mapping_review.xlsx"
SUMMARY = ROOT / "work_logs" / "b02_p0_review_summary.md"


P0_HINTS = {
    "SYS-001": ["江湖人生", "职业等级", "不是DND", "门派只作为背景", "不是职业"],
    "SYS-002": ["纸笔可跑", "APP", "辅助"],
    "CHK-001": ["开卡", "角色创建", "角色卡", "九步", "背景包"],
    "ATT-001": ["筋骨", "身法", "根基", "心法", "眼力", "气势"],
    "SKL-001": ["熟练", "入门", "精通", "大师", "宗师", "d4", "d6", "d8", "d10", "d12"],
    "SLOT-001": ["命中槽"],
    "SLOT-002": ["得利槽"],
    "SLOT-003": ["代价槽", "破绽槽", "破绽"],
    "SLOT-004": ["得利不能", "万能兑换", "授权", "规则菜单", "装备", "关键词"],
    "SCN-001": ["场景进度", "危机"],
    "SCN-002": ["核心线索", "必定获得", "必给"],
    "SCN-003": ["线索深度", "提前量", "代价"],
    "SCN-004": ["调查", "交涉", "潜入", "追逐"],
    "SCN-005": ["失败仍推进", "失败不堵", "失败推进"],
    "SCN-006": ["高代价成功", "代价成功", "话术"],
    "CBT-001": ["伤害结算", "伤害公式", "伤害"],
    "CBT-002": ["势态", "守势", "攻势"],
    "CBT-003": ["Pos", "身位"],
    "CBT-004": ["破绽"],
    "CBT-005": ["普通敌人", "破绽兑现", "削得利", "退势", "轻反应"],
    "CBT-006": ["精英", "Boss", "反击层级", "高手"],
    "CBT-007": ["反击链", "不触发新的反击", "终止"],
    "ACT-001": ["协助行动", "协助"],
    "ACT-002": ["强行尝试", "我能不能", "临场发挥"],
    "ACT-003": ["移动", "位移"],
    "ACT-004": ["非杀伤", "制服"],
    "ACT-005": ["环境互动", "环境"],
    "MA-001": ["博览", "全精", "不可全精"],
    "MA-002": ["主修1门", "主修一门", "主修"],
    "MA-003": ["兼修1-2门", "兼修"],
    "MA-004": ["旁学若干", "旁学"],
    "NG-001": ["主修内功", "辅修", "主修1", "辅修1"],
    "NG-002": ["旁学内功", "不叠", "不叠机制"],
    "GROW-001": ["感悟", "成长", "经验升级"],
    "FAME-001": ["名望0", "名望", "0-5"],
    "FAME-002": ["名望", "世界权限"],
    "FAME-003": ["名望调用", "调用入口"],
    "REST-001": ["三阶段休整", "休整"],
    "REST-002": ["岁月事件", "岁月"],
    "REST-003": ["年龄节点", "线性惩罚"],
    "FAC-001": ["世界投放链", "世界投放"],
    "NPC-001": ["NPC生成", "NPC"],
    "NPC-002": ["敌人分层", "普通敌人", "精英", "Boss"],
    "ADV-001": ["剧本生成", "剧本"],
    "CARD-001": ["武器卡"],
    "CARD-002": ["招式卡"],
    "CARD-003": ["内功卡"],
    "REC-001": ["三轨记录", "DM三轨", "记录表"],
    "EX-001": ["得利强度表", "示例"],
}


def normalize(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def load_p0_rules() -> list[dict[str, str]]:
    wb = load_workbook(EXEC_PACK, data_only=True)
    ws = wb.worksheets[3]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and row[2] == "P0":
            rows.append({"rule_id": str(row[0]), "rule_name": str(row[1]), "priority": "P0"})
    return rows


def load_chunks() -> list[dict[str, object]]:
    with CHUNKS_JSONL.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def load_mapped_counts() -> dict[str, int]:
    counts: dict[str, int] = {}
    with MAPPED_JSONL.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            rid = str(obj.get("mapped_rule_id", ""))
            if "NEW?" in rid:
                continue
            counts[rid] = counts.get(rid, 0) + 1
    return counts


def chunk_score(rule: dict[str, str], chunk: dict[str, object]) -> int:
    text = normalize(str(chunk.get("chapter_or_heading", "")) + "\n" + str(chunk.get("text", "")))
    score = 0
    for hint in P0_HINTS.get(rule["rule_id"], []):
        hint_norm = normalize(hint)
        if hint_norm and hint_norm in text:
            score += 3
    for token in re.split(r"[：:/、，,（）()]+", rule["rule_name"]):
        token_norm = normalize(token)
        if len(token_norm) >= 2 and token_norm in text:
            score += 2
    module = rule["rule_id"].split("-")[0]
    if module in str(chunk.get("chunk_id", "")):
        score += 1
    return score


def excerpt_for(rule: dict[str, str], chunk: dict[str, object]) -> str:
    text = str(chunk.get("text", ""))
    hints = P0_HINTS.get(rule["rule_id"], [])
    best_pos = None
    for hint in hints:
        pos = text.find(hint)
        if pos >= 0:
            best_pos = pos
            break
    if best_pos is None:
        return text[:220].replace("\n", " ")
    start = max(0, best_pos - 80)
    end = min(len(text), best_pos + 180)
    return text[start:end].replace("\n", " ")


def review_rules(rules: list[dict[str, str]], chunks: list[dict[str, object]], mapped_counts: dict[str, int]) -> list[dict[str, object]]:
    reviews = []
    for rule in rules:
        scored = sorted(((chunk_score(rule, chunk), chunk) for chunk in chunks), key=lambda item: (-item[0], str(item[1].get("chunk_id"))))
        top_score, top_chunk = scored[0]
        previous_hits = mapped_counts.get(rule["rule_id"], 0)
        if previous_hits > 0 and top_score >= 3:
            status = "已覆盖-有候选和回查证据"
        elif previous_hits > 0:
            status = "已覆盖-仅候选命中"
        elif top_score >= 6:
            status = "补命中-需加入映射"
        elif top_score >= 3:
            status = "疑似命中-需人工复核"
        else:
            status = "疑似缺失-需人工确认"
        reviews.append(
            {
                "rule_id": rule["rule_id"],
                "rule_name": rule["rule_name"],
                "priority": rule["priority"],
                "previous_candidate_hits": previous_hits,
                "review_status": status,
                "evidence_score": top_score,
                "evidence_chunk_id": top_chunk.get("chunk_id"),
                "source_type": top_chunk.get("source_type"),
                "source_file": top_chunk.get("source_file"),
                "chapter_or_heading": top_chunk.get("chapter_or_heading"),
                "position": top_chunk.get("position"),
                "excerpt": excerpt_for(rule, top_chunk),
                "limitations": "半自动回查；需要B02人工复核后再进入B03。岁月数据库缺失时REST/LIFE相关规则为限制性结论。",
            }
        )
    return reviews


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P0复核结果"
    headers = [
        "Rule_ID",
        "规则名",
        "优先级",
        "首版候选命中数",
        "复核状态",
        "证据分数",
        "证据chunk",
        "来源类型",
        "章节/标题",
        "位置",
        "证据摘录",
        "来源文件",
        "限制说明",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["rule_id"],
                row["rule_name"],
                row["priority"],
                row["previous_candidate_hits"],
                row["review_status"],
                row["evidence_score"],
                row["evidence_chunk_id"],
                row["source_type"],
                row["chapter_or_heading"],
                json.dumps(row["position"], ensure_ascii=False),
                row["excerpt"],
                row["source_file"],
                row["limitations"],
            ]
        )
    fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 28, 10, 14, 22, 10, 18, 14, 26, 24, 70, 48, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_summary(rows: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["review_status"]] = counts.get(row["review_status"], 0) + 1
    lines = [
        "# B02 P0 Rule_ID 定向复核摘要",
        "",
        "生成时间：2026-06-11 03:03 Asia/Shanghai",
        "",
        "## 输出",
        "",
        f"- JSONL：`{OUT_JSONL}`",
        f"- XLSX：`{OUT_XLSX}`",
        f"- P0 Rule_ID总数：{len(rows)}",
        "",
        "## 复核状态分布",
        "",
        "| 状态 | 数量 |",
        "| --- | ---: |",
    ]
    for key, value in sorted(counts.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## 需要优先人工处理", ""])
    urgent = [row for row in rows if row["review_status"] in {"疑似缺失-需人工确认", "疑似命中-需人工复核"}]
    if urgent:
        for row in urgent:
            lines.append(f"- `{row['rule_id']}` {row['rule_name']}：{row['review_status']}，证据分数 {row['evidence_score']}，chunk `{row['evidence_chunk_id']}`")
    else:
        lines.append("- 无")
    lines.extend(
        [
            "",
            "## 限制",
            "",
            "- 本轮是P0定向回查，不直接修改规则书。",
            "- `补命中-需加入映射` 表示规则书文本中有疑似证据，但上一轮候选映射未覆盖，需要下一轮合并进03映射表。",
            "- `疑似缺失-需人工确认` 不能直接判定规则缺失，需人工读证据块后再进入B03冲突/缺失审查。",
            "- 岁月与生成数据库仍缺失，REST/LIFE相关结论继续限制性处理。",
        ]
    )
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rules = load_p0_rules()
    chunks = load_chunks()
    mapped_counts = load_mapped_counts()
    rows = review_rules(rules, chunks, mapped_counts)
    write_jsonl(rows)
    write_xlsx(rows)
    write_summary(rows)
    print(json.dumps({"p0_rules": len(rules), "review_rows": len(rows), "jsonl": str(OUT_JSONL), "xlsx": str(OUT_XLSX)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
EXEC_PACK = ROOT / "_审查执行包_v0.42.1_解压" / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
INPUT_JSONL = ROOT / "review_output" / "02_rules_extracted.jsonl"
OUT_JSONL = ROOT / "review_output" / "03_rules_mapped.jsonl"
OUT_XLSX = ROOT / "review_output" / "03_rules_mapped.xlsx"
SUMMARY = ROOT / "work_logs" / "b02_rule_mapping_summary.md"


RULE_HINTS = {
    "SYS-001": ["江湖人生", "职业等级", "不是DND", "门派", "背景"],
    "SYS-002": ["纸笔", "APP", "辅助", "记录", "查询"],
    "CHK-001": ["开卡", "角色创建", "角色卡", "出身", "背景包"],
    "ATT-001": ["筋骨", "身法", "根基", "心法", "眼力", "气势", "六属性"],
    "SKL-001": ["熟练", "入门", "精通", "大师", "宗师", "d4", "d6", "d8", "d10", "d12"],
    "SLOT-001": ["命中槽", "命中"],
    "SLOT-002": ["得利槽", "得利"],
    "SLOT-003": ["代价槽", "代价", "破绽槽", "破绽"],
    "SLOT-004": ["得利不能", "万能兑换", "授权", "菜单"],
    "SCN-001": ["场景进度", "危机", "场景"],
    "SCN-002": ["核心线索", "必定获得", "必给"],
    "SCN-003": ["线索深度", "提前量", "代价"],
    "SCN-004": ["调查", "交涉", "潜入", "追逐"],
    "COM-001": ["战斗", "回合", "行动"],
    "COM-002": ["势态", "守势", "攻势"],
    "COM-003": ["Pos", "身位", "位置"],
    "COM-004": ["反击", "反击链"],
    "COM-005": ["普通敌人", "精英", "Boss", "高手"],
    "MART-001": ["武学容量", "主修", "兼修", "旁学"],
    "MART-002": ["内功容量", "主修内功", "辅修", "旁学内功"],
    "MART-003": ["轻功", "移动", "追逐"],
    "WPN-001": ["武器", "兵器"],
    "WPN-002": ["护甲", "防具"],
    "RES-001": ["名望", "世界权限"],
    "RES-002": ["人情"],
    "RES-003": ["情报"],
    "RES-004": ["把柄"],
    "LIFE-001": ["休整"],
    "LIFE-002": ["年龄", "节点"],
    "LIFE-003": ["家庭", "弟子", "产业"],
    "DM-001": ["DC", "难度"],
    "DM-002": ["情景骰", "S骰", "S "],
    "DM-003": ["NPC", "敌人", "操作"],
    "DB-001": ["数据库", "字段", "索引"],
}


def load_rule_index() -> list[dict[str, str]]:
    wb = load_workbook(EXEC_PACK, data_only=True)
    ws = wb.worksheets[3]
    rules: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        rules.append(
            {
                "rule_id": str(row[0]),
                "rule_name": "" if row[1] is None else str(row[1]),
                "priority": "" if row[2] is None else str(row[2]),
            }
        )
    return rules


def load_candidates() -> list[dict[str, object]]:
    candidates: list[dict[str, object]] = []
    with INPUT_JSONL.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                candidates.append(json.loads(line))
    return candidates


def normalize(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def score_rule(candidate: dict[str, object], rule: dict[str, str]) -> int:
    text = normalize(" ".join(str(candidate.get(key, "")) for key in ["excerpt", "rule_name", "chapter_or_heading", "suggested_module"]))
    rid = rule["rule_id"]
    score = 0
    module = rid.split("-")[0]
    if str(candidate.get("suggested_module")) == module:
        score += 2
    for hint in RULE_HINTS.get(rid, []):
        if normalize(hint) in text:
            score += 3
    for token in re.split(r"[：:/、，,（）()]+", rule["rule_name"]):
        token = normalize(token)
        if len(token) >= 2 and token in text:
            score += 2
    if rule["priority"] == "P0" and score:
        score += 1
    return score


def map_candidate(candidate: dict[str, object], rules: list[dict[str, str]]) -> dict[str, object]:
    scored = sorted(((score_rule(candidate, rule), rule) for rule in rules), key=lambda item: (-item[0], item[1]["rule_id"]))
    best_score, best_rule = scored[0]
    if best_score >= 7:
        status = "已映射-高置信"
    elif best_score >= 4:
        status = "已映射-需复核"
    else:
        best_rule = {"rule_id": f"{candidate.get('suggested_module', 'MISC')}-NEW?", "rule_name": "待新增或待人工映射", "priority": "待定"}
        status = "未映射-待人工确认"
    mapped = dict(candidate)
    mapped.update(
        {
            "mapped_rule_id": best_rule["rule_id"],
            "mapped_rule_name": best_rule["rule_name"],
            "mapped_priority": best_rule["priority"],
            "mapping_score": best_score,
            "mapping_status": status,
            "mapping_note": "半自动关键词映射；B02后续需人工复核低置信和未映射项。",
        }
    )
    return mapped


def write_jsonl(items: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def write_xlsx(items: list[dict[str, object]], rules: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "规则映射结果"
    headers = [
        "候选编号",
        "映射Rule_ID",
        "映射规则名",
        "优先级",
        "映射状态",
        "映射分数",
        "来源类型",
        "章节/标题",
        "位置",
        "原文摘录",
        "输入/触发",
        "输出/结算",
        "影响资源",
        "原建议模块",
        "原建议Rule_ID",
        "来源文件",
        "备注",
    ]
    ws.append(headers)
    for item in items:
        ws.append(
            [
                item.get("candidate_id"),
                item.get("mapped_rule_id"),
                item.get("mapped_rule_name"),
                item.get("mapped_priority"),
                item.get("mapping_status"),
                item.get("mapping_score"),
                item.get("source_type"),
                item.get("chapter_or_heading"),
                item.get("position"),
                item.get("excerpt"),
                item.get("input_trigger"),
                item.get("resolution"),
                item.get("affected_resources"),
                item.get("suggested_module"),
                item.get("suggested_rule_id"),
                item.get("source_file"),
                item.get("mapping_note"),
            ]
        )
    fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 16, 28, 10, 18, 10, 14, 24, 22, 62, 30, 34, 20, 12, 18, 46, 38]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Rule_ID覆盖摘要")
    ws2.append(["Rule_ID", "规则名", "优先级", "映射候选数", "高置信", "需复核"])
    for rule in rules:
        subset = [item for item in items if item.get("mapped_rule_id") == rule["rule_id"]]
        high = sum(1 for item in subset if item.get("mapping_status") == "已映射-高置信")
        review = sum(1 for item in subset if item.get("mapping_status") == "已映射-需复核")
        ws2.append([rule["rule_id"], rule["rule_name"], rule["priority"], len(subset), high, review])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws2.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_summary(items: list[dict[str, object]], rules: list[dict[str, str]]) -> None:
    status_counts: dict[str, int] = {}
    rule_counts: dict[str, int] = {}
    for item in items:
        status_counts[str(item["mapping_status"])] = status_counts.get(str(item["mapping_status"]), 0) + 1
        rid = str(item["mapped_rule_id"])
        rule_counts[rid] = rule_counts.get(rid, 0) + 1
    covered_p0 = [rule for rule in rules if rule["priority"] == "P0" and rule_counts.get(rule["rule_id"], 0) > 0]
    all_p0 = [rule for rule in rules if rule["priority"] == "P0"]
    uncovered_p0 = [rule for rule in all_p0 if rule_counts.get(rule["rule_id"], 0) == 0]
    top_rules = sorted(rule_counts.items(), key=lambda item: (-item[1], item[0]))[:20]

    lines = [
        "# B02 Rule_ID 映射摘要",
        "",
        "生成时间：2026-06-11 02:33 Asia/Shanghai",
        "",
        "## 输出",
        "",
        f"- JSONL：`{OUT_JSONL}`",
        f"- XLSX：`{OUT_XLSX}`",
        f"- 映射候选总数：{len(items)}",
        f"- 核心Rule_ID数量：{len(rules)}",
        "",
        "## 映射状态",
        "",
        "| 状态 | 数量 |",
        "| --- | ---: |",
    ]
    for key, value in sorted(status_counts.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(
        [
            "",
            "## P0覆盖",
            "",
            f"- P0 Rule_ID总数：{len(all_p0)}",
            f"- 已命中P0 Rule_ID：{len(covered_p0)}",
            f"- 未命中P0 Rule_ID：{len(uncovered_p0)}",
            "",
            "### 未命中P0 Rule_ID",
            "",
        ]
    )
    if uncovered_p0:
        for rule in uncovered_p0:
            lines.append(f"- `{rule['rule_id']}` {rule['rule_name']}")
    else:
        lines.append("- 无")
    lines.extend(["", "## 映射最多的Rule_ID", "", "| Rule_ID | 候选数 |", "| --- | ---: |"])
    for rid, count in top_rules:
        lines.append(f"| {rid} | {count} |")
    lines.extend(
        [
            "",
            "## 限制",
            "",
            "- 本轮为半自动关键词映射，不等于最终规则审查结论。",
            "- `未映射-待人工确认` 与 `已映射-需复核` 必须在后续B02复核或B03冲突审查中处理。",
            "- CHK/SKL/SLOT等P0模块若未命中或命中过少，需人工回查规则书切块。",
            "- 岁月与生成数据库仍缺失，LIFE相关映射只能支撑限制性结论。",
        ]
    )
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rules = load_rule_index()
    candidates = load_candidates()
    mapped = [map_candidate(candidate, rules) for candidate in candidates]
    write_jsonl(mapped)
    write_xlsx(mapped, rules)
    write_summary(mapped, rules)
    print(json.dumps({"rules": len(rules), "candidates": len(candidates), "mapped": len(mapped), "jsonl": str(OUT_JSONL), "xlsx": str(OUT_XLSX)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

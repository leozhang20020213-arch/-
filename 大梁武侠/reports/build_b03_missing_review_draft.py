from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
INPUT = ROOT / "review_output" / "04_b03_missing_review_input.jsonl"
OUT_JSONL = ROOT / "review_output" / "04_b03_missing_review_draft.jsonl"
OUT_XLSX = ROOT / "review_output" / "04_b03_missing_review_draft.xlsx"
OUT_MD = ROOT / "review_output" / "04_b03_missing_review_draft.md"
SUMMARY = ROOT / "work_logs" / "b03_missing_review_draft_summary.md"


IMPACT = {
    "ACT": "玩家临场行动入口不清，普通玩家容易问“我能不能这样做”，DM需要临场补裁定。",
    "CARD": "上桌引用介质不足，玩家和DM难以快速调用招式/装备/内功信息。",
    "REC": "DM记录负担高，危机、解密、战役得利等轨道容易漏记。",
    "EX": "示例不足会导致得利强度和兑换边界难理解，容易误用为万能资源。",
    "CBT": "战斗流程证据不足，反击链、伤害或破绽可能导致等待、混乱或无限触发。",
    "SLOT": "三槽边界不清会影响核心检定理解。",
    "SYS": "游戏定位证据不足会影响整书口径。",
    "REST": "休整/年龄相关结论受岁月数据库缺失影响，只能限制性判断。",
    "FAME": "名望调用入口不清会让资源像货币或摆设，影响社交玩法。",
    "FAC": "世界投放链证据不足会影响长团投放和剧本生成。",
}


def load_rows() -> list[dict[str, object]]:
    with INPUT.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def classify(row: dict[str, object]) -> dict[str, object]:
    rid = str(row["rule_id"])
    module = rid.split("-")[0]
    triage = str(row["triage_status"])
    if "确认疑似缺失" in triage:
        review_result = "疑似缺失"
        handling = "优先补示例/入口；若B03后续确认正文无规则，再列入修订方案。"
    else:
        review_result = "弱证据需复核"
        handling = "人工阅读证据块；能支持则补入覆盖表，不能支持再转疑似缺失。"
    if rid in {"REC-001", "EX-001", "ACT-002", "ACT-004"}:
        severity = "P0"
    elif module in {"CARD", "ACT", "CBT"}:
        severity = "P0/P1"
    else:
        severity = "P1"
    limitation = "岁月数据库缺失不直接影响本项。"
    if module in {"REST", "FAC"}:
        limitation = "受岁月与生成数据库缺失影响，结论为限制性。"
    item = dict(row)
    item.update(
        {
            "review_result": review_result,
            "severity": severity,
            "impact": IMPACT.get(module, "需在B03后续审查中补充影响说明。"),
            "handling_recommendation": handling,
            "may_modify_now": "否",
            "limitation_note": limitation,
        }
    )
    return item


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "B03缺失审查草案"
    headers = ["Rule_ID", "规则名", "审查结论", "严重度", "原分类", "证据分数", "证据chunk", "影响", "处理建议", "可否立即改", "摘录", "来源文件", "限制"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["rule_id"],
            row["rule_name"],
            row["review_result"],
            row["severity"],
            row["triage_status"],
            row["evidence_score"],
            row["evidence_chunk_id"],
            row["impact"],
            row["handling_recommendation"],
            row["may_modify_now"],
            row["excerpt"],
            row["source_file"],
            row["limitation_note"],
        ])
    fill = PatternFill("solid", fgColor="FCE4D6")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 30, 16, 10, 24, 10, 18, 52, 52, 12, 70, 48, 42]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    counts = {}
    for row in rows:
        counts[row["review_result"]] = counts.get(row["review_result"], 0) + 1
    lines = [
        "# B03 缺失/弱证据审查草案",
        "",
        "生成时间：2026-06-11 05:03 Asia/Shanghai",
        "",
        "## 结论边界",
        "",
        "本文件是B03草案，不是规则修改方案；所有条目只列证据、影响和建议处理方式，不直接修改规则书。",
        "",
        "## 统计",
        "",
        "| 结论 | 数量 |",
        "| --- | ---: |",
    ]
    for key, value in sorted(counts.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## 条目", ""])
    for row in rows:
        lines.extend([
            f"### `{row['rule_id']}` {row['rule_name']}",
            "",
            f"- 审查结论：{row['review_result']}",
            f"- 严重度：{row['severity']}",
            f"- 证据：`{row['evidence_chunk_id']}`，分数 {row['evidence_score']}",
            f"- 影响：{row['impact']}",
            f"- 建议：{row['handling_recommendation']}",
            f"- 限制：{row['limitation_note']}",
            "",
        ])
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    missing = sum(1 for row in rows if row["review_result"] == "疑似缺失")
    weak = sum(1 for row in rows if row["review_result"] == "弱证据需复核")
    lines = [
        "# B03 缺失审查草案摘要",
        "",
        "生成时间：2026-06-11 05:03 Asia/Shanghai",
        "",
        f"- 输入项：{len(rows)}",
        f"- 疑似缺失：{missing}",
        f"- 弱证据需复核：{weak}",
        f"- JSONL：`{OUT_JSONL}`",
        f"- XLSX：`{OUT_XLSX}`",
        f"- Markdown：`{OUT_MD}`",
        "",
        "限制：本轮不改规则书；岁月数据库仍缺失。",
    ]
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = [classify(row) for row in load_rows()]
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"items": len(rows), "jsonl": str(OUT_JSONL), "xlsx": str(OUT_XLSX), "md": str(OUT_MD)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

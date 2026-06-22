from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
EXEC_PACK = ROOT / "_审查执行包_v0.42.1_解压" / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
DB_COVERAGE = ROOT / "review_output" / "06_db_coverage_draft.jsonl"
OUT_JSONL = ROOT / "review_output" / "07_math_test_plan_draft.jsonl"
OUT_XLSX = ROOT / "review_output" / "07_math_test_plan.xlsx"
OUT_MD = ROOT / "review_output" / "07_math_test_plan_draft.md"
SUMMARY = ROOT / "work_logs" / "b06_math_test_plan_draft_summary.md"


def load_matrix() -> list[dict[str, object]]:
    wb = load_workbook(EXEC_PACK, data_only=True)
    ws = wb.worksheets[6]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        rows.append(
            {
                "test_id": str(row[0]),
                "topic": str(row[1] or ""),
                "target": str(row[2] or ""),
                "variables": str(row[3] or ""),
                "sample_size": str(row[4] or ""),
                "metrics": str(row[5] or ""),
                "standard": str(row[6] or ""),
                "priority": str(row[7] or ""),
            }
        )
    return rows


def load_db_risks() -> set[str]:
    risks = set()
    if not DB_COVERAGE.exists():
        return risks
    with DB_COVERAGE.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            if obj.get("draft_result") != "规则与数据均有证据-待复核":
                risks.add(str(obj.get("domain", "")))
    return risks


def infer_method(row: dict[str, object], db_risks: set[str]) -> tuple[str, str, str]:
    text = f"{row['topic']} {row['target']} {row['variables']}"
    if any(k in text for k in ["武学", "内功", "敌人", "战斗", "伤害"]):
        method = "蒙特卡洛战斗模拟"
        data_need = "武学/内功/敌人/装备数据库字段"
    elif any(k in text for k in ["休整", "岁月", "年龄", "家庭", "弟子", "产业"]):
        method = "长团资源流模拟"
        data_need = "岁月与生成数据库、休整/年龄/家庭/产业表"
    elif any(k in text for k in ["资源", "名望", "人情", "情报", "把柄"]):
        method = "资源经济压力测试"
        data_need = "资源入口、消耗、恢复与场景奖励表"
    else:
        method = "参数扫描与边界测试"
        data_need = "规则映射表与相关数据库字段"
    limitation = "无阻塞"
    if any(domain in data_need for domain in db_risks):
        limitation = "数据库解释存在风险，测试结果需标限制性"
    return method, data_need, limitation


def build_rows() -> list[dict[str, object]]:
    db_risks = load_db_risks()
    rows = []
    for item in load_matrix():
        method, data_need, limitation = infer_method(item, db_risks)
        executable = "可执行草案"
        if "待填" in item["target"] or not item["target"]:
            executable = "需补测试对象"
        rows.append(
            {
                **item,
                "method": method,
                "data_need": data_need,
                "execution_status": executable,
                "output_artifact": f"math_results/{item['test_id']}.jsonl",
                "limitation": limitation,
                "may_modify_now": "否",
            }
        )
    return rows


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "数值测试计划"
    headers = [
        "测试ID", "测试主题", "测试对象", "变量控制", "样本规模建议", "关键指标", "判定标准",
        "优先级", "建议方法", "数据需求", "执行状态", "输出文件", "限制", "可否立即改"
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["test_id"], row["topic"], row["target"], row["variables"], row["sample_size"], row["metrics"],
            row["standard"], row["priority"], row["method"], row["data_need"], row["execution_status"],
            row["output_artifact"], row["limitation"], row["may_modify_now"],
        ])
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 24, 34, 34, 18, 34, 44, 10, 24, 44, 16, 28, 32, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    p0 = [r for r in rows if r["priority"] == "P0"]
    lines = ["# B06 数值测试计划草案", "", "生成时间：2026-06-11 10:33 Asia/Shanghai", "", f"- 测试项：{len(rows)}", f"- P0测试项：{len(p0)}", "", "## P0测试项", ""]
    for row in p0:
        lines.append(f"- `{row['test_id']}` {row['topic']}：{row['method']}，输出 `{row['output_artifact']}`")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    methods: dict[str, int] = {}
    for row in rows:
        methods[row["method"]] = methods.get(row["method"], 0) + 1
    lines = ["# B06 数值测试计划草案摘要", "", "生成时间：2026-06-11 10:33 Asia/Shanghai", "", f"- 测试项：{len(rows)}", f"- P0测试项：{sum(1 for r in rows if r['priority']=='P0')}"]
    for key, value in sorted(methods.items()):
        lines.append(f"- {key}：{value}")
    lines.extend([f"- JSONL：`{OUT_JSONL}`", f"- XLSX：`{OUT_XLSX}`", f"- Markdown：`{OUT_MD}`", "", "限制：本轮只生成测试计划，不运行数值模拟，不修改规则书。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"math_tests": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

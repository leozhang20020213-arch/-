from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
EXEC_PACK = ROOT / "_审查执行包_v0.42.1_解压" / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
OUT_JSONL = ROOT / "review_output" / "08_playtest_plan_draft.jsonl"
OUT_XLSX = ROOT / "review_output" / "08_playtest_plan.xlsx"
OUT_MD = ROOT / "review_output" / "08_playtest_plan_draft.md"
SUMMARY = ROOT / "work_logs" / "b07_playtest_plan_draft_summary.md"


PLAYER_ROLES = ["player_newbie", "player_story", "player_combat", "player_builder", "player_casual"]


def load_matrix() -> list[dict[str, object]]:
    wb = load_workbook(EXEC_PACK, data_only=True)
    ws = wb.worksheets[7]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        rows.append(
            {
                "test_id": str(row[0]),
                "test_type": str(row[1] or ""),
                "scenario": str(row[2] or ""),
                "players": str(row[3] or ""),
                "dm_requirement": str(row[4] or ""),
                "focus": str(row[5] or ""),
                "priority": str(row[6] or ""),
            }
        )
    return rows


def assign_roles(row: dict[str, object]) -> str:
    text = f"{row['test_type']} {row['scenario']} {row['players']} {row['focus']}"
    roles = ["player_newbie", "player_story", "player_casual"]
    if any(k in text for k in ["战斗", "护镖", "边关", "团战", "武器", "Boss"]):
        roles.append("player_combat")
    if any(k in text for k in ["成长", "装备", "武学", "内功", "构筑", "资源"]):
        roles.append("player_builder")
    for role in PLAYER_ROLES:
        if role not in roles:
            roles.append(role)
    return "；".join(roles[:5])


def infer_output(row: dict[str, object]) -> tuple[str, str]:
    tid = row["test_id"]
    log = f"playtest_results/{tid}_session_log.md"
    sheet = f"playtest_results/{tid}_feedback.xlsx"
    return log, sheet


def build_rows() -> list[dict[str, object]]:
    rows = []
    for item in load_matrix():
        log, sheet = infer_output(item)
        rows.append(
            {
                **item,
                "agent_roster": assign_roles(item),
                "dm_agent": "dm",
                "supervisor": "supervisor_editor",
                "feedback_form": "普通玩家反馈表 + DM负担记录 + 监管修改触发条件",
                "session_log_output": log,
                "feedback_output": sheet,
                "execution_status": "可执行草案",
                "may_modify_now": "否",
                "limitation": "本轮只生成跑团压力测试计划；不进行实跑，不修改规则书。",
            }
        )
    return rows


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "跑团压力测试计划"
    headers = [
        "测试ID", "测试类型", "模拟内容", "玩家配置", "DM要求", "重点观察", "优先级",
        "玩家智能体", "DM智能体", "监管智能体", "反馈格式", "日志输出", "反馈输出", "状态", "可否立即改", "限制"
    ]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["test_id"], row["test_type"], row["scenario"], row["players"], row["dm_requirement"],
            row["focus"], row["priority"], row["agent_roster"], row["dm_agent"], row["supervisor"],
            row["feedback_form"], row["session_log_output"], row["feedback_output"], row["execution_status"],
            row["may_modify_now"], row["limitation"],
        ])
    fill = PatternFill("solid", fgColor="EADCF8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 24, 40, 28, 34, 40, 10, 46, 14, 18, 42, 34, 34, 16, 12, 44]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    p0 = [r for r in rows if r["priority"] == "P0"]
    lines = ["# B07 跑团压力测试计划草案", "", "生成时间：2026-06-11 10:58 Asia/Shanghai", "", f"- 测试项：{len(rows)}", f"- P0测试项：{len(p0)}", "", "## P0测试项", ""]
    for row in p0:
        lines.append(f"- `{row['test_id']}` {row['test_type']}：{row['scenario']}；观察：{row['focus']}")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    priorities: dict[str, int] = {}
    for row in rows:
        priorities[row["priority"]] = priorities.get(row["priority"], 0) + 1
    lines = ["# B07 跑团压力测试计划草案摘要", "", "生成时间：2026-06-11 10:58 Asia/Shanghai", "", f"- 测试项：{len(rows)}"]
    for key, value in sorted(priorities.items()):
        lines.append(f"- {key}：{value}")
    lines.extend([f"- JSONL：`{OUT_JSONL}`", f"- XLSX：`{OUT_XLSX}`", f"- Markdown：`{OUT_MD}`", "", "限制：本轮只生成跑团压力测试计划，不进行实跑，不修改规则书。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"playtest_items": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

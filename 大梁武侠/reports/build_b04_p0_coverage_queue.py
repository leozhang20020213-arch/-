from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
INPUT = ROOT / "review_output" / "05_coverage_draft.jsonl"
OUT_JSONL = ROOT / "review_output" / "05_b04_p0_coverage_queue.jsonl"
OUT_XLSX = ROOT / "review_output" / "05_b04_p0_coverage_queue.xlsx"
OUT_MD = ROOT / "review_output" / "05_b04_p0_coverage_queue.md"
SUMMARY = ROOT / "work_logs" / "b04_p0_coverage_queue_summary.md"


def load_rows() -> list[dict[str, object]]:
    with INPUT.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def risk_score(row: dict[str, object]) -> int:
    result = str(row["draft_result"])
    score = 0
    if result == "覆盖不足-待确认":
        score += 5
    elif result == "需复核":
        score += 4
    elif result == "部分覆盖-需补证据":
        score += 3
    obj = str(row.get("object", "")) + str(row.get("standard", ""))
    for key in ["P0", "触发", "输出", "示例", "卡片", "数据库", "得利", "行动", "记录", "规则"]:
        if key in obj:
            score += 1
    if int(row.get("evidence_score", 0)) <= 2:
        score += 2
    return score


def recommend(row: dict[str, object]) -> str:
    text = str(row.get("object", "")) + str(row.get("scope", "")) + str(row.get("standard", ""))
    if "示例" in text:
        return "优先补示例，不直接改核心规则。"
    if "卡片" in text:
        return "优先补卡片/速查介质。"
    if "数据库" in text:
        return "优先做数据库引用核对，确认是规则缺数据还是数据缺解释。"
    if "触发" in text or "输出" in text:
        return "优先补规则入口/触发/结算说明。"
    if row["draft_result"] == "覆盖不足-待确认":
        return "进入覆盖缺失审查，先人工确认证据。"
    return "补证据并人工复核。"


def build_queue(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    queue = []
    for row in rows:
        if row.get("priority") == "P0" and row.get("draft_result") != "基本覆盖-待人工确认":
            queue.append({
                **row,
                "risk_score": risk_score(row),
                "next_action": recommend(row),
                "may_modify_now": "否",
            })
    return sorted(queue, key=lambda row: (-int(row["risk_score"]), str(row["check_id"])))


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P0覆盖风险队列"
    headers = ["排序", "检查ID", "对象", "范围", "标准", "草案结果", "风险分", "证据分", "证据来源", "证据位置", "建议动作", "可否立即改", "摘录", "来源文件", "限制"]
    ws.append(headers)
    for idx, row in enumerate(rows, 1):
        ws.append([
            idx, row["check_id"], row["object"], row["scope"], row["standard"], row["draft_result"],
            row["risk_score"], row["evidence_score"], row["evidence_source"],
            f"{row['evidence_chunk_id']} {row['position']}", row["next_action"], row["may_modify_now"],
            row["excerpt"], row["source_file"], row["limitations"],
        ])
    fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [8, 12, 30, 34, 46, 20, 10, 10, 24, 24, 42, 12, 70, 48, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    lines = ["# B04 P0覆盖风险队列", "", "生成时间：2026-06-11 07:03 Asia/Shanghai", "", "本队列只用于优先级排序，不直接修改规则书。", ""]
    for idx, row in enumerate(rows[:15], 1):
        lines.extend([
            f"## {idx}. `{row['check_id']}` {row['object']}",
            f"- 结果：{row['draft_result']}",
            f"- 风险分：{row['risk_score']}",
            f"- 建议动作：{row['next_action']}",
            "",
        ])
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    actions: dict[str, int] = {}
    for row in rows:
        counts[str(row["draft_result"])] = counts.get(str(row["draft_result"]), 0) + 1
        actions[str(row["next_action"])] = actions.get(str(row["next_action"]), 0) + 1
    lines = ["# B04 P0覆盖风险队列摘要", "", "生成时间：2026-06-11 07:03 Asia/Shanghai", "", f"- 队列项：{len(rows)}"]
    for key, value in sorted(counts.items()):
        lines.append(f"- {key}：{value}")
    lines.append("")
    lines.append("## 建议动作分布")
    for key, value in sorted(actions.items()):
        lines.append(f"- {key}：{value}")
    lines.extend([f"- JSONL：`{OUT_JSONL}`", f"- XLSX：`{OUT_XLSX}`", f"- Markdown：`{OUT_MD}`", "", "限制：本轮只排队，不修改规则书。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_queue(load_rows())
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"queue": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

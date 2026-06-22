from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
INPUT = ROOT / "review_output" / "04_conflicts_draft.jsonl"
OUT_JSONL = ROOT / "review_output" / "04_b03_p0_conflict_queue.jsonl"
OUT_XLSX = ROOT / "review_output" / "04_b03_p0_conflict_queue.xlsx"
OUT_MD = ROOT / "review_output" / "04_b03_p0_conflict_queue.md"
SUMMARY = ROOT / "work_logs" / "b03_p0_conflict_queue_summary.md"


HIGH_RISK_TOPICS = ["反击", "破绽", "得利", "内功", "武学", "熟练", "伤害", "休整", "年龄", "记录"]


def load_rows() -> list[dict[str, object]]:
    with INPUT.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def risk_score(row: dict[str, object]) -> int:
    score = 0
    if row.get("draft_result") == "缺失-待确认":
        score += 4
    if row.get("draft_result") == "需人工判断":
        score += 3
    topic = str(row.get("topic", "")) + str(row.get("keywords", "")) + str(row.get("final_position", ""))
    for key in HIGH_RISK_TOPICS:
        if key in topic:
            score += 2
    if int(row.get("evidence_score", 0)) == 0:
        score += 2
    return score


def build_queue(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    queue = [
        {
            **row,
            "queue_reason": "P0且草案结果为缺失待确认/需人工判断，不能直接进入修订。",
            "risk_score": risk_score(row),
            "next_action": "人工读取证据块与相关章节；确认最终结果为通过/冲突/缺失/需改之一。",
            "may_modify_now": "否",
        }
        for row in rows
        if row.get("priority") == "P0" and row.get("draft_result") in {"缺失-待确认", "需人工判断"}
    ]
    return sorted(queue, key=lambda row: (-int(row["risk_score"]), str(row["conflict_id"])))


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P0冲突复核队列"
    headers = ["排序", "冲突ID", "主题", "草案结果", "风险分", "关键词", "最终口径", "证据chunk", "证据分数", "摘录", "下一动作", "可否立即改", "来源文件", "限制"]
    ws.append(headers)
    for idx, row in enumerate(rows, 1):
        ws.append([
            idx,
            row.get("conflict_id"),
            row.get("topic"),
            row.get("draft_result"),
            row.get("risk_score"),
            row.get("keywords"),
            row.get("final_position"),
            row.get("evidence_chunk_id"),
            row.get("evidence_score"),
            row.get("excerpt"),
            row.get("next_action"),
            row.get("may_modify_now"),
            row.get("source_file"),
            row.get("limitations"),
        ])
    fill = PatternFill("solid", fgColor="F4CCCC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [8, 12, 24, 18, 10, 42, 46, 18, 10, 70, 48, 12, 48, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    lines = [
        "# B03 P0冲突复核队列",
        "",
        "生成时间：2026-06-11 06:03 Asia/Shanghai",
        "",
        "本队列只用于人工/监管智能体复核，不是修订方案。所有条目均不可立即改规则书。",
        "",
    ]
    for idx, row in enumerate(rows, 1):
        lines.extend([
            f"## {idx}. `{row['conflict_id']}` {row['topic']}",
            "",
            f"- 草案结果：{row['draft_result']}",
            f"- 风险分：{row['risk_score']}",
            f"- 最终口径：{row['final_position']}",
            f"- 证据：`{row['evidence_chunk_id']}`，分数 {row['evidence_score']}",
            f"- 下一动作：{row['next_action']}",
            "",
        ])
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        counts[str(row["draft_result"])] = counts.get(str(row["draft_result"]), 0) + 1
    lines = [
        "# B03 P0冲突复核队列摘要",
        "",
        "生成时间：2026-06-11 06:03 Asia/Shanghai",
        "",
        f"- 队列项：{len(rows)}",
    ]
    for key, value in sorted(counts.items()):
        lines.append(f"- {key}：{value}")
    lines.extend([
        f"- JSONL：`{OUT_JSONL}`",
        f"- XLSX：`{OUT_XLSX}`",
        f"- Markdown：`{OUT_MD}`",
        "",
        "限制：本轮只排队，不裁定最终冲突，不修改规则书。",
    ])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_queue(load_rows())
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"queue": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
DB_FILES = [
    ("基础数据库", ROOT / "大梁武侠TRPG_数据库填充_V1.xlsx"),
    ("岁月与生成数据库", ROOT / "大梁武侠TRPG_岁月与生成数据库_重生成版_v1.1.xlsx"),
    ("世界投放数据库", ROOT / "大梁武侠TRPG_世界投放与剧本生成数据库_V1.xlsx"),
    ("总库索引", ROOT / "大梁武侠TRPG_总数据库索引与规则书成稿计划_V1.xlsx"),
]
MAPPING = ROOT / "review_output" / "03_rules_mapped_final_b02.jsonl"
OUT_JSONL = ROOT / "review_output" / "06_db_coverage_draft.jsonl"
OUT_XLSX = ROOT / "review_output" / "06_db_coverage.xlsx"
OUT_MD = ROOT / "review_output" / "06_db_coverage_draft.md"
SUMMARY = ROOT / "work_logs" / "b05_db_coverage_draft_summary.md"

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "office_rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

EXPECTED_DOMAINS = {
    "武学": ["武学", "招式", "武功"],
    "内功": ["内功"],
    "轻功": ["轻功"],
    "装备": ["装备", "武器", "护甲", "兵器"],
    "背景": ["背景", "出身"],
    "敌人/NPC": ["敌人", "NPC", "Boss", "高手"],
    "世界投放": ["世界", "投放", "地图", "秘境", "势力"],
    "剧本生成": ["剧本", "任务", "事件"],
    "岁月/年龄": ["岁月", "年龄", "年岁", "节点"],
    "家庭/弟子/产业": ["家庭", "弟子", "产业"],
    "休整": ["休整", "疗伤", "修行"],
    "总索引": ["索引", "总库", "计划"],
}

RULE_TO_DOMAIN = {
    "MART": ["武学", "内功"],
    "MA": ["武学"],
    "NG": ["内功"],
    "WPN": ["装备"],
    "CBT": ["敌人/NPC", "装备"],
    "NPC": ["敌人/NPC"],
    "ADV": ["剧本生成"],
    "FAC": ["世界投放"],
    "REST": ["休整", "岁月/年龄", "家庭/弟子/产业"],
    "LIFE": ["岁月/年龄", "家庭/弟子/产业"],
    "FAME": ["世界投放"],
    "CARD": ["装备", "武学", "内功"],
}


def norm(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        data = zf.read("xl/sharedStrings.xml")
    except KeyError:
        return []
    root = ET.fromstring(data)
    strings = []
    for si in root.findall("main:si", NS):
        pieces = [t.text or "" for t in si.findall(".//main:t", NS)]
        strings.append("".join(pieces))
    return strings


def workbook_sheets(path: Path) -> list[dict[str, object]]:
    with zipfile.ZipFile(path) as zf:
        strings = shared_strings(zf)
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall("rel:Relationship", NS)}
        sheets = []
        for sheet in wb.findall("main:sheets/main:sheet", NS):
            name = sheet.attrib["name"]
            rid = sheet.attrib[f"{{{NS['office_rel']}}}id"]
            target = rel_map[rid].lstrip("/")
            xml_path = "xl/" + target if not target.startswith("xl/") else target
            try:
                ws_xml = ET.fromstring(zf.read(xml_path))
            except KeyError:
                sheets.append({"sheet": name, "dimension": "", "headers": []})
                continue
            dim_node = ws_xml.find("main:dimension", NS)
            dimension = "" if dim_node is None else dim_node.attrib.get("ref", "")
            headers = []
            for row in ws_xml.findall("main:sheetData/main:row", NS)[:3]:
                values = []
                for cell in row.findall("main:c", NS)[:12]:
                    value_node = cell.find("main:v", NS)
                    if value_node is None:
                        inline = cell.find(".//main:t", NS)
                        values.append("" if inline is None else inline.text or "")
                        continue
                    value = value_node.text or ""
                    if cell.attrib.get("t") == "s":
                        try:
                            value = strings[int(value)]
                        except Exception:
                            pass
                    values.append(value)
                if any(values):
                    headers = values
                    break
            sheets.append({"sheet": name, "dimension": dimension, "headers": headers})
        return sheets


def load_mapped_modules() -> dict[str, int]:
    counts: dict[str, int] = {}
    with MAPPING.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            rid = str(obj.get("mapped_rule_id", ""))
            module = rid.split("-")[0] if "-" in rid else str(obj.get("suggested_module", "MISC"))
            counts[module] = counts.get(module, 0) + 1
    return counts


def detect_domains(inventory: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    found: dict[str, list[dict[str, object]]] = {domain: [] for domain in EXPECTED_DOMAINS}
    for item in inventory:
        haystack = norm(str(item["workbook"]) + str(item["sheet"]) + " ".join(map(str, item.get("headers", []))))
        for domain, keys in EXPECTED_DOMAINS.items():
            if any(norm(key) in haystack for key in keys):
                found[domain].append(item)
    return found


def build_rows() -> list[dict[str, object]]:
    inventory: list[dict[str, object]] = []
    for label, path in DB_FILES:
        if not path.exists():
            inventory.append({"workbook": label, "file": str(path), "sheet": "", "dimension": "", "headers": [], "status": "文件缺失"})
            continue
        for sheet in workbook_sheets(path):
            inventory.append({"workbook": label, "file": str(path), **sheet, "status": "已读取"})
    domains = detect_domains(inventory)
    module_counts = load_mapped_modules()

    rows = []
    for domain, evidence in domains.items():
        related_modules = [m for m, ds in RULE_TO_DOMAIN.items() if domain in ds]
        rule_hits = sum(module_counts.get(m, 0) for m in related_modules)
        if evidence and rule_hits:
            result = "规则与数据均有证据-待复核"
            action = "核对字段是否足以支撑规则触发/结算。"
        elif evidence and not rule_hits:
            result = "数据有证据-规则解释不足"
            action = "检查规则书是否解释该数据库用途。"
        elif not evidence and rule_hits:
            result = "规则有证据-数据支撑不足"
            action = "检查是否缺少数据库表或字段。"
        else:
            result = "规则与数据均不足-待确认"
            action = "人工确认是否属于当前版本范围。"
        rows.append({
            "domain": domain,
            "draft_result": result,
            "related_rule_modules": "；".join(related_modules),
            "mapped_rule_hits": rule_hits,
            "database_evidence_count": len(evidence),
            "database_evidence": "；".join(f"{e['workbook']}::{e['sheet']}[{e['dimension']}]" for e in evidence[:5]),
            "headers_sample": " | ".join("、".join(map(str, e.get("headers", [])[:6])) for e in evidence[:3]),
            "handling_note": action,
            "may_modify_now": "否",
            "limitations": "数据库覆盖率草案；只判断证据存在性，不判定字段语义完全正确。",
        })
    return rows


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "数据库覆盖率草案"
    headers = ["领域", "草案结果", "关联规则模块", "映射规则命中", "数据库证据数", "数据库证据", "表头样例", "处理建议", "可否立即改", "限制"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in ["domain", "draft_result", "related_rule_modules", "mapped_rule_hits", "database_evidence_count", "database_evidence", "headers_sample", "handling_note", "may_modify_now", "limitations"]])
    fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [18, 26, 24, 14, 14, 60, 70, 42, 12, 48]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    lines = ["# B05 数据库覆盖率草案", "", "生成时间：2026-06-11 09:35 Asia/Shanghai", "", "| 领域 | 结果 | 数据证据 | 规则命中 |", "| --- | --- | ---: | ---: |"]
    for row in rows:
        lines.append(f"| {row['domain']} | {row['draft_result']} | {row['database_evidence_count']} | {row['mapped_rule_hits']} |")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["draft_result"]] = counts.get(row["draft_result"], 0) + 1
    lines = ["# B05 数据库覆盖率草案摘要", "", "生成时间：2026-06-11 09:35 Asia/Shanghai", "", f"- 检查领域：{len(rows)}"]
    for key, value in sorted(counts.items()):
        lines.append(f"- {key}：{value}")
    lines.extend([f"- JSONL：`{OUT_JSONL}`", f"- XLSX：`{OUT_XLSX}`", f"- Markdown：`{OUT_MD}`", "", "说明：岁月与生成数据库已找到，本轮解除此前输入缺失阻塞；本草案不修改规则书。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"db_domains": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

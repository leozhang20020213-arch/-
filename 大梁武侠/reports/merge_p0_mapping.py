from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
MAPPED_JSONL = ROOT / "review_output" / "03_rules_mapped.jsonl"
P0_REVIEW_JSONL = ROOT / "review_output" / "03_p0_rule_mapping_review.jsonl"
OUT_MERGED_JSONL = ROOT / "review_output" / "03_rules_mapped_p0_merged.jsonl"
OUT_MERGED_XLSX = ROOT / "review_output" / "03_rules_mapped_p0_merged.xlsx"
OUT_MANUAL_JSONL = ROOT / "review_output" / "03_p0_manual_review_queue.jsonl"
OUT_MANUAL_XLSX = ROOT / "review_output" / "03_p0_manual_review_queue.xlsx"
SUMMARY = ROOT / "work_logs" / "b02_merge_p0_summary.md"


def load_jsonl(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def synthetic_candidate(row: dict[str, object], index: int) -> dict[str, object]:
    rid = str(row["rule_id"])
    module = rid.split("-")[0]
    return {
        "candidate_id": f"P0FIX-{index:03d}",
        "source_file": row["source_file"],
        "source_type": row["source_type"],
        "chunk_id": row["evidence_chunk_id"],
        "chapter_or_heading": row["chapter_or_heading"],
        "position": json.dumps(row["position"], ensure_ascii=False),
        "suggested_module": module,
        "suggested_rule_id": rid,
        "rule_name": row["rule_name"],
        "excerpt": row["excerpt"],
        "input_trigger": "P0定向回查补命中，待人工确认触发条件",
        "resolution": "P0定向回查补命中，待人工确认结算方式",
        "affected_resources": "待人工确认",
        "database_refs": "待人工确认",
        "ambiguity": "需复核",
        "status": "P0补命中",
        "score": row["evidence_score"],
        "limitations": row["limitations"],
        "mapped_rule_id": rid,
        "mapped_rule_name": row["rule_name"],
        "mapped_priority": row["priority"],
        "mapping_score": row["evidence_score"],
        "mapping_status": "P0补命中-待复核",
        "mapping_note": "由B02-P0-REVIEW从章节块回查补入；进入B03前需人工确认摘录是否真正支持该Rule_ID。",
    }


def write_mapping_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P0合并映射结果"
    headers = [
        "候选编号",
        "映射Rule_ID",
        "映射规则名",
        "优先级",
        "映射状态",
        "映射分数",
        "来源类型",
        "章节/标题",
        "位置",
        "原文摘录",
        "输入/触发",
        "输出/结算",
        "影响资源",
        "原建议模块",
        "来源文件",
        "备注",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.get("candidate_id"),
                row.get("mapped_rule_id"),
                row.get("mapped_rule_name"),
                row.get("mapped_priority"),
                row.get("mapping_status"),
                row.get("mapping_score"),
                row.get("source_type"),
                row.get("chapter_or_heading"),
                row.get("position"),
                row.get("excerpt"),
                row.get("input_trigger"),
                row.get("resolution"),
                row.get("affected_resources"),
                row.get("suggested_module"),
                row.get("source_file"),
                row.get("mapping_note"),
            ]
        )
    fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 16, 30, 10, 22, 10, 14, 26, 24, 70, 32, 36, 22, 14, 48, 46]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(path)


def write_manual_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "P0人工处理清单"
    headers = [
        "Rule_ID",
        "规则名",
        "处理状态",
        "证据分数",
        "证据chunk",
        "来源类型",
        "章节/标题",
        "位置",
        "证据摘录",
        "建议处理",
        "来源文件",
        "限制说明",
    ]
    ws.append(headers)
    for row in rows:
        if row["review_status"] == "疑似命中-需人工复核":
            suggestion = "人工阅读证据块；若摘录支持规则，则补入映射；否则转为疑似缺失。"
        else:
            suggestion = "人工全文检索并核查；若仍无证据，进入B03缺失审查，不得直接改稿。"
        ws.append(
            [
                row["rule_id"],
                row["rule_name"],
                row["review_status"],
                row["evidence_score"],
                row["evidence_chunk_id"],
                row["source_type"],
                row["chapter_or_heading"],
                json.dumps(row["position"], ensure_ascii=False),
                row["excerpt"],
                suggestion,
                row["source_file"],
                row["limitations"],
            ]
        )
    fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 30, 22, 10, 18, 14, 26, 24, 70, 46, 48, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    mapped = load_jsonl(MAPPED_JSONL)
    p0_review = load_jsonl(P0_REVIEW_JSONL)
    fixes = [row for row in p0_review if row["review_status"] == "补命中-需加入映射"]
    manual = [row for row in p0_review if row["review_status"] in {"疑似命中-需人工复核", "疑似缺失-需人工确认"}]
    synthetic = [synthetic_candidate(row, index + 1) for index, row in enumerate(fixes)]
    merged = mapped + synthetic

    write_jsonl(OUT_MERGED_JSONL, merged)
    write_jsonl(OUT_MANUAL_JSONL, manual)
    write_mapping_xlsx(OUT_MERGED_XLSX, merged)
    write_manual_xlsx(OUT_MANUAL_XLSX, manual)

    summary_lines = [
        "# B02 P0补命中合并摘要",
        "",
        "生成时间：2026-06-11 03:33 Asia/Shanghai",
        "",
        "## 输出",
        "",
        f"- 合并映射JSONL：`{OUT_MERGED_JSONL}`",
        f"- 合并映射XLSX：`{OUT_MERGED_XLSX}`",
        f"- 人工处理JSONL：`{OUT_MANUAL_JSONL}`",
        f"- 人工处理XLSX：`{OUT_MANUAL_XLSX}`",
        "",
        "## 结果",
        "",
        f"- 原映射候选：{len(mapped)}",
        f"- P0补命中合并：{len(synthetic)}",
        f"- 合并后映射候选：{len(merged)}",
        f"- 人工处理项：{len(manual)}",
        "",
        "## 人工处理项",
        "",
    ]
    for row in manual:
        summary_lines.append(f"- `{row['rule_id']}` {row['rule_name']}：{row['review_status']}，chunk `{row['evidence_chunk_id']}`")
    summary_lines.extend(
        [
            "",
            "## 限制",
            "",
            "- 本轮不覆盖原始 `03_rules_mapped`，只生成合并版。",
            "- P0补命中仍标记为待复核，进入B03前应人工确认摘录是否支持对应Rule_ID。",
            "- 18条人工处理项未解决前，B03冲突审查只能作为草案执行。",
            "- 岁月与生成数据库仍缺失，REST/LIFE相关结论继续限制性处理。",
        ]
    )
    SUMMARY.write_text("\n".join(summary_lines) + "\n", encoding="utf-8")

    print(json.dumps({"original": len(mapped), "p0_fixes": len(synthetic), "merged": len(merged), "manual": len(manual)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

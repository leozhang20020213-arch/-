from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"D:\trpg\大梁武侠")
UNPACKED = ROOT / "_审查执行包_v0.42.1_解压"
OUT_DIR = ROOT / "reports"
RAW_DIR = OUT_DIR / "raw_logs"

PACK_XLSX = UNPACKED / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
REPORT_MD = OUT_DIR / "v0.42.1规则工程化审查执行包_审查报告.md"
RAW_JSON = RAW_DIR / "v0421_execution_pack_audit.json"


REQUIRED_INPUTS = [
    ("玩家规则书", "大梁武侠TRPG_玩家规则书_完整整合版_v0.41.docx"),
    ("DM规则书", "大梁武侠TRPG_DM规则书_完整整合版_v0.41.docx"),
    ("基础数据库", "大梁武侠TRPG_数据库填充_V1.xlsx"),
    ("岁月与生成数据库", "大梁武侠TRPG_岁月与生成数据库_重生成版_v1.1.xlsx"),
    ("世界投放数据库", "大梁武侠TRPG_世界投放与剧本生成数据库_V1.xlsx"),
    ("总库索引", "大梁武侠TRPG_总数据库索引与规则书成稿计划_V1.xlsx"),
    ("执行包工作簿", "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"),
]


def all_project_files() -> list[Path]:
    ignored = {".git", "__pycache__"}
    files: list[Path] = []
    for path in ROOT.rglob("*"):
        if not path.is_file():
            continue
        if any(part in ignored for part in path.parts):
            continue
        files.append(path)
    return files


def find_by_name(files: list[Path], name: str) -> list[Path]:
    return [p for p in files if p.name == name]


def sheet_summary(wb) -> list[dict[str, object]]:
    summaries: list[dict[str, object]] = []
    marker_terms = ["待执行", "未开始", "待填", "Codex填写", "待抽取", "未检查"]
    for ws in wb.worksheets:
        markers = {term: 0 for term in marker_terms}
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if not isinstance(value, str):
                    continue
                for term in marker_terms:
                    if term in value:
                        markers[term] += 1
        summaries.append(
            {
                "sheet": ws.title,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "markers": {k: v for k, v in markers.items() if v},
            }
        )
    return summaries


def first_data_rows(wb, limit: int = 5) -> dict[str, list[list[str]]]:
    result: dict[str, list[list[str]]] = {}
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, limit), values_only=True):
            rows.append(["" if value is None else str(value) for value in row])
        result[ws.title] = rows
    return result


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    RAW_DIR.mkdir(exist_ok=True)

    project_files = all_project_files()
    inventory = []
    for label, filename in REQUIRED_INPUTS:
        matches = find_by_name(project_files, filename)
        inventory.append(
            {
                "label": label,
                "filename": filename,
                "status": "found" if matches else "missing",
                "paths": [str(p) for p in matches],
            }
        )

    wb = load_workbook(PACK_XLSX, data_only=True)
    sheets = sheet_summary(wb)
    rows = first_data_rows(wb)

    audit = {
        "pack": str(PACK_XLSX),
        "sheet_count": len(wb.worksheets),
        "sheets": sheets,
        "required_inputs": inventory,
        "sample_rows": rows,
        "conclusion": {
            "is_execution_pack": True,
            "is_completed_audit": False,
            "main_blocker": "岁月与生成数据库_重生成版_v1.1.xlsx 未在当前项目文件中找到；工作簿结果栏仍为待执行/待填。",
        },
    }

    RAW_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")

    missing = [item for item in inventory if item["status"] == "missing"]
    found = [item for item in inventory if item["status"] == "found"]

    def fmt_paths(paths: list[str]) -> str:
        return "；".join(paths) if paths else "未找到"

    md: list[str] = []
    md.append("# v0.42.1 规则工程化审查执行包审查报告\n")
    md.append("## 结论\n")
    md.append(
        "v0.42.1 已从 v0.42 的“主控审查框架”升级为可交给 Codex 执行的审查作业包：它补充了输入风险表、全文规则抽取导入表、Rule_ID 映射表、冲突/覆盖/数学/跑团矩阵和执行提示词。"
    )
    md.append(
        "但它本身仍不是已完成审查结果。工作簿中大量单元格仍是“待执行”“待填”“Codex填写”，所以不能把它当作最终规则审查结论或 v0.42 修订方案。"
    )
    md.append("\n## 输入文件核对\n")
    md.append("| 文件类型 | 状态 | 路径 |")
    md.append("| --- | --- | --- |")
    for item in inventory:
        md.append(f"| {item['label']} | {item['status']} | {fmt_paths(item['paths'])} |")

    md.append("\n## 工作簿结构\n")
    md.append("| 表名 | 行数 | 列数 | 待填/待执行标记 |")
    md.append("| --- | ---: | ---: | --- |")
    for sheet in sheets:
        markers = ", ".join(f"{k}:{v}" for k, v in sheet["markers"].items()) or "未检出"
        md.append(f"| {sheet['sheet']} | {sheet['rows']} | {sheet['cols']} | {markers} |")

    md.append("\n## 对 v0.42 缺口的承接情况\n")
    md.append("| v0.42 主要缺口 | v0.42.1 是否承接 | 审查判断 |")
    md.append("| --- | --- | --- |")
    md.append("| 输入文件风险未显式化 | 是 | 新增 `输入文件风险表`，并在提示词中要求缺文件先报告。当前项目仍缺岁月与生成数据库。 |")
    md.append("| 全文规则抽取缺执行入口 | 是 | 新增 150 条预留抽取表，但内容尚未抽取。 |")
    md.append("| Rule_ID 与正文位置未映射 | 是 | 新增 48 条核心映射表，但结果仍待执行。 |")
    md.append("| 冲突审查结果栏待填 | 部分承接 | 扩展为 30 项执行表，要求证据和原文摘录，但尚未执行。 |")
    md.append("| 覆盖率审查不完整 | 部分承接 | 扩展为 35 项执行表，覆盖正文、示例、卡片、数据库引用，但尚未执行。 |")
    md.append("| 数值与跑团矩阵不足 | 是 | 数值矩阵 35 项、跑团压力矩阵 30 项，范围明显扩展。 |")
    md.append("| 最终修订方案缺失 | 否 | 执行提示词要求输出 `10_final_review.md` 和 `11_v042_patch_plan.xlsx`，但包内尚未包含这些结果。 |")

    md.append("\n## 关键问题\n")
    md.append("1. 当前包是“执行模板”，不是“执行结果”。必须跑完提示词中的 11 个输出文件，才算完成审查。")
    md.append("2. `岁月与生成数据库_重生成版_v1.1.xlsx` 当前未找到，长团、年龄、家庭、弟子、产业相关验证会失真。")
    md.append("3. 玩家书和 DM 书在子目录中，提示词写的是同一目录查找；实际执行器应递归查找，或先整理输入目录。")
    md.append("4. `全文规则抽取导入表` 预留 150 条规则候选，但没有自动抽取结果；下一步必须从 docx 正文生成真实候选。")
    md.append("5. 执行包要求数值测试计划，但不等于已完成数值模拟；如果要判断超模，还需要实际模拟脚本和结果。")

    md.append("\n## 建议下一步\n")
    md.append("1. 先补齐或确认缺失的岁月与生成数据库；若确实没有，应在输入风险表标为缺失并限制长团结论范围。")
    md.append("2. 按执行提示词生成 `review_output/00_file_inventory.md` 到 `review_output/11_v042_patch_plan.xlsx`。")
    md.append("3. 审查完成后，再让 `supervisor_editor` 根据多玩家反馈、DM 卡点和数值结果决定修改/补示例/暂缓。")
    md.append("4. 不建议直接拿 v0.42.1 工作簿改规则书；它还缺证据摘录、Rule_ID 映射和实际测试结果。")

    md.append("\n## 原始记录\n")
    md.append(f"- JSON 审查记录：`{RAW_JSON}`")

    REPORT_MD.write_text("\n".join(md) + "\n", encoding="utf-8")
    print(REPORT_MD)
    print(RAW_JSON)


if __name__ == "__main__":
    main()

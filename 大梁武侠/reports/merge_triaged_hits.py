from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
BASE = ROOT / "review_output" / "03_rules_mapped_p0_merged.jsonl"
TRIAGE = ROOT / "review_output" / "03_p0_manual_queue_triaged.jsonl"
OUT_FINAL_JSONL = ROOT / "review_output" / "03_rules_mapped_final_b02.jsonl"
OUT_FINAL_XLSX = ROOT / "review_output" / "03_rules_mapped_final_b02.xlsx"
OUT_B03_INPUT_JSONL = ROOT / "review_output" / "04_b03_missing_review_input.jsonl"
OUT_B03_INPUT_XLSX = ROOT / "review_output" / "04_b03_missing_review_input.xlsx"
SUMMARY = ROOT / "work_logs" / "b02_merge_triaged_hits_summary.md"


def load_jsonl(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def write_jsonl(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def make_hit(row: dict[str, object], index: int) -> dict[str, object]:
    rid = str(row["rule_id"])
    module = rid.split("-")[0]
    return {
        "candidate_id": f"TRIAGEFIX-{index:03d}",
        "source_file": row["source_file"],
        "source_type": row["source_type"],
        "chunk_id": row["evidence_chunk_id"],
        "chapter_or_heading": row["chapter_or_heading"],
        "position": json.dumps(row["position"], ensure_ascii=False),
        "suggested_module": module,
        "suggested_rule_id": rid,
        "rule_name": row["rule_name"],
        "excerpt": row["excerpt"],
        "input_trigger": "人工队列分类确认命中，待B03审查触发条件",
        "resolution": "人工队列分类确认命中，待B03审查结算方式",
        "affected_resources": "待B03确认",
        "database_refs": "待B03确认",
        "ambiguity": "需B03复核",
        "status": "人工队列确认命中",
        "score": row["evidence_score"],
        "limitations": row["limitations"],
        "mapped_rule_id": rid,
        "mapped_rule_name": row["rule_name"],
        "mapped_priority": "P0",
        "mapping_score": row["evidence_score"],
        "mapping_status": "人工队列确认命中-待B03复核",
        "mapping_note": "由B02-MANUAL-QUEUE-TRIAGE确认命中后补入；B03仍需引用证据判断冲突/覆盖。",
    }


def write_mapping_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "B02最终映射"
    headers = ["候选编号", "Rule_ID", "规则名", "优先级", "状态", "分数", "来源", "章节", "位置", "摘录", "触发", "结算", "资源", "来源文件", "备注"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row.get("candidate_id"),
            row.get("mapped_rule_id"),
            row.get("mapped_rule_name"),
            row.get("mapped_priority"),
            row.get("mapping_status"),
            row.get("mapping_score"),
            row.get("source_type"),
            row.get("chapter_or_heading"),
            row.get("position"),
            row.get("excerpt"),
            row.get("input_trigger"),
            row.get("resolution"),
            row.get("affected_resources"),
            row.get("source_file"),
            row.get("mapping_note"),
        ])
    fill = PatternFill("solid", fgColor="E2F0D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [16, 16, 30, 10, 28, 10, 12, 24, 24, 70, 34, 34, 20, 48, 46]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(path)


def write_b03_xlsx(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "B03缺失审查输入"
    headers = ["Rule_ID", "规则名", "分类", "证据分数", "证据chunk", "来源", "章节", "位置", "摘录", "B03建议动作", "来源文件", "限制"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["rule_id"],
            row["rule_name"],
            row["triage_status"],
            row["evidence_score"],
            row["evidence_chunk_id"],
            row["source_type"],
            row["chapter_or_heading"],
            json.dumps(row["position"], ensure_ascii=False),
            row["excerpt"],
            "作为B03缺失/覆盖审查输入；先列证据与影响，不直接改规则。",
            row["source_file"],
            row["limitations"],
        ])
    fill = PatternFill("solid", fgColor="F4CCCC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 30, 24, 10, 18, 12, 24, 24, 70, 46, 48, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    base = load_jsonl(BASE)
    triaged = load_jsonl(TRIAGE)
    hits = [row for row in triaged if row["triage_status"] == "确认命中-可补映射"]
    weak = [row for row in triaged if row["triage_status"] == "继续待查-有弱证据"]
    missing = [row for row in triaged if row["triage_status"] == "确认疑似缺失-进B03缺失审查"]
    additions = [make_hit(row, i + 1) for i, row in enumerate(hits)]
    final = base + additions
    b03_input = weak + missing

    write_jsonl(OUT_FINAL_JSONL, final)
    write_jsonl(OUT_B03_INPUT_JSONL, b03_input)
    write_mapping_xlsx(OUT_FINAL_XLSX, final)
    write_b03_xlsx(OUT_B03_INPUT_XLSX, b03_input)

    lines = [
        "# B02 确认命中合并摘要",
        "",
        "生成时间：2026-06-11 04:33 Asia/Shanghai",
        "",
        "## 输出",
        "",
        f"- B02最终映射JSONL：`{OUT_FINAL_JSONL}`",
        f"- B02最终映射XLSX：`{OUT_FINAL_XLSX}`",
        f"- B03缺失审查输入JSONL：`{OUT_B03_INPUT_JSONL}`",
        f"- B03缺失审查输入XLSX：`{OUT_B03_INPUT_XLSX}`",
        "",
        "## 结果",
        "",
        f"- B02合并前映射：{len(base)}",
        f"- 本轮确认命中补入：{len(additions)}",
        f"- B02最终映射：{len(final)}",
        f"- B03输入项：{len(b03_input)}，其中弱证据 {len(weak)}，疑似缺失 {len(missing)}",
        "",
        "## 限制",
        "",
        "- B02最终映射仍不是规则修改方案，只是B03冲突/覆盖审查输入。",
        "- B03输入项应先审查证据和影响，不得直接改规则。",
        "- 岁月与生成数据库仍缺失，REST/FAC相关结论继续限制性处理。",
    ]
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(json.dumps({"base": len(base), "added_hits": len(additions), "final": len(final), "b03_input": len(b03_input), "weak": len(weak), "missing": len(missing)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

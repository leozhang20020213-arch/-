from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
INPUT_JSONL = ROOT / "review_output" / "01_chapter_chunks.jsonl"
OUT_JSONL = ROOT / "review_output" / "02_rules_extracted.jsonl"
OUT_XLSX = ROOT / "review_output" / "02_rules_extracted.xlsx"
SUMMARY = ROOT / "work_logs" / "b01_rule_candidates_summary.md"


MODULE_KEYWORDS = [
    ("SYS", ["纸笔", "APP", "江湖人生", "定位", "规则冲突"]),
    ("CHK", ["开卡", "角色创建", "角色卡", "背景", "出身"]),
    ("ATT", ["筋骨", "身法", "根基", "心法", "眼力", "气势", "属性"]),
    ("SKL", ["熟练", "入门", "精通", "大师", "宗师", "d4", "d6", "d8", "d10", "d12"]),
    ("SLOT", ["命中槽", "得利槽", "代价槽", "三槽", "分槽"]),
    ("SCN", ["情景", "场景", "进度", "危机", "解密", "调查", "交涉", "潜入", "追逐"]),
    ("CUE", ["线索", "证据", "核心线索", "线索深度"]),
    ("COM", ["战斗", "回合", "先攻", "伤害", "反击", "破绽", "势态", "Pos"]),
    ("WPN", ["武器", "护甲", "装备", "兵器"]),
    ("MART", ["武学", "招式", "连招", "内功", "轻功", "主修", "兼修", "旁学"]),
    ("RES", ["得利", "名望", "人情", "情报", "把柄", "资源"]),
    ("DM", ["DM", "主持", "裁定", "DC", "NPC", "敌人", "Boss"]),
    ("LIFE", ["休整", "岁月", "年龄", "家庭", "弟子", "产业", "长团"]),
    ("DB", ["数据库", "表", "字段", "索引"]),
]

RULE_TERMS = [
    "必须",
    "不得",
    "不能",
    "可以",
    "可",
    "需要",
    "消耗",
    "获得",
    "增加",
    "减少",
    "触发",
    "判定",
    "检定",
    "结算",
    "记录",
    "选择",
    "声明",
    "失败",
    "成功",
    "代价",
    "奖励",
    "若",
    "当",
    "每",
    "默认",
]


def load_chunks() -> list[dict[str, object]]:
    chunks: list[dict[str, object]] = []
    with INPUT_JSONL.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                chunks.append(json.loads(line))
    return chunks


def split_sentences(text: str) -> list[str]:
    raw = re.split(r"(?<=[。！？；;])\s*|\n+", text)
    return [s.strip() for s in raw if len(s.strip()) >= 8]


def score_sentence(sentence: str) -> int:
    score = 0
    for term in RULE_TERMS:
        if term in sentence:
            score += 2
    for _, keywords in MODULE_KEYWORDS:
        for keyword in keywords:
            if keyword in sentence:
                score += 1
    if re.search(r"\d+|d\d+", sentence, flags=re.I):
        score += 1
    if len(sentence) > 180:
        score -= 1
    return score


def infer_module(text: str) -> str:
    scores: dict[str, int] = {}
    for module, keywords in MODULE_KEYWORDS:
        for keyword in keywords:
            if keyword in text:
                scores[module] = scores.get(module, 0) + 1
    if not scores:
        return "MISC"
    return sorted(scores.items(), key=lambda item: (-item[1], item[0]))[0][0]


def infer_trigger(sentence: str) -> str:
    patterns = [
        r"(当[^，。；;]{2,40})",
        r"(若[^，。；;]{2,40})",
        r"(如果[^，。；;]{2,40})",
        r"(每当[^，。；;]{2,40})",
        r"(在[^，。；;]{2,40}时)",
    ]
    for pattern in patterns:
        match = re.search(pattern, sentence)
        if match:
            return match.group(1)
    if "检定" in sentence or "判定" in sentence:
        return "进行检定/判定时"
    if "战斗" in sentence:
        return "战斗流程中"
    if "情景" in sentence or "场景" in sentence:
        return "情景/场景流程中"
    return "待人工确认"


def infer_resolution(sentence: str) -> str:
    for marker in ["则", "可以", "必须", "获得", "消耗", "增加", "减少", "记录", "结算"]:
        idx = sentence.find(marker)
        if idx >= 0:
            return sentence[idx : idx + 80]
    return "待人工确认"


def affected_resources(sentence: str) -> str:
    resources = []
    for key in ["命中槽", "得利槽", "代价槽", "得利", "破绽", "势态", "Pos", "名望", "人情", "情报", "把柄", "危机", "解密", "场景进度"]:
        if key in sentence:
            resources.append(key)
    return "；".join(resources)


def extract_candidates(chunks: list[dict[str, object]]) -> list[dict[str, object]]:
    candidates: list[dict[str, object]] = []
    seen: set[tuple[str, str]] = set()
    for chunk in chunks:
        text = str(chunk["text"])
        sentences = split_sentences(text)
        ranked = sorted(
            [(score_sentence(sentence), sentence) for sentence in sentences],
            key=lambda item: (-item[0], len(item[1])),
        )
        for score, sentence in ranked[:4]:
            if score < 4:
                continue
            key = (str(chunk["chunk_id"]), sentence)
            if key in seen:
                continue
            seen.add(key)
            module = infer_module(sentence + " " + str(chunk["chapter_or_heading"]))
            candidates.append(
                {
                    "candidate_id": f"EXT-{len(candidates) + 1:04d}",
                    "source_file": chunk["source_file"],
                    "source_type": chunk["source_type"],
                    "chunk_id": chunk["chunk_id"],
                    "chapter_or_heading": chunk["chapter_or_heading"],
                    "position": json.dumps(chunk["position"], ensure_ascii=False),
                    "suggested_module": module,
                    "suggested_rule_id": f"{module}-待映射",
                    "rule_name": sentence[:30],
                    "excerpt": sentence,
                    "input_trigger": infer_trigger(sentence),
                    "resolution": infer_resolution(sentence),
                    "affected_resources": affected_resources(sentence),
                    "database_refs": "待人工确认",
                    "ambiguity": "待人工确认",
                    "status": "待映射",
                    "score": score,
                    "limitations": chunk.get("limitations", ""),
                }
            )
    return candidates


def write_jsonl(candidates: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for item in candidates:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def write_xlsx(candidates: list[dict[str, object]]) -> None:
    headers = [
        "候选编号",
        "来源文件",
        "来源类型",
        "章节/标题",
        "页码/位置",
        "建议模块",
        "建议Rule_ID",
        "规则名称",
        "规则原文摘录",
        "输入/触发",
        "输出/结算",
        "影响资源",
        "关联数据库",
        "是否歧义",
        "处理状态",
        "抽取分数",
        "限制说明",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "全文规则抽取导入表"
    ws.append(headers)
    for item in candidates:
        ws.append(
            [
                item["candidate_id"],
                item["source_file"],
                item["source_type"],
                item["chapter_or_heading"],
                item["position"],
                item["suggested_module"],
                item["suggested_rule_id"],
                item["rule_name"],
                item["excerpt"],
                item["input_trigger"],
                item["resolution"],
                item["affected_resources"],
                item["database_refs"],
                item["ambiguity"],
                item["status"],
                item["score"],
                item["limitations"],
            ]
        )
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 46, 14, 24, 20, 12, 18, 30, 60, 28, 32, 24, 18, 14, 14, 10, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_summary(candidates: list[dict[str, object]]) -> None:
    by_source: dict[str, int] = {}
    by_module: dict[str, int] = {}
    for item in candidates:
        by_source[item["source_type"]] = by_source.get(item["source_type"], 0) + 1
        by_module[item["suggested_module"]] = by_module.get(item["suggested_module"], 0) + 1
    lines = [
        "# B01 规则候选抽取摘要",
        "",
        "生成时间：2026-06-11 02:03 Asia/Shanghai",
        "",
        "## 输出",
        "",
        f"- JSONL：`{OUT_JSONL}`",
        f"- XLSX：`{OUT_XLSX}`",
        f"- 候选总数：{len(candidates)}",
        "",
        "## 来源分布",
        "",
        "| 来源 | 候选数 |",
        "| --- | ---: |",
    ]
    for key, value in sorted(by_source.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## 模块分布", "", "| 模块 | 候选数 |", "| --- | ---: |"])
    for key, value in sorted(by_module.items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"| {key} | {value} |")
    lines.extend(
        [
            "",
            "## 限制",
            "",
            "- 本轮是启发式规则候选抽取，不等于最终 Rule_ID 映射。",
            "- `是否歧义`、`关联数据库`、`建议Rule_ID` 仍需 B02 人工/半自动映射确认。",
            "- 候选来源位置沿用 `01_chapter_chunks.jsonl` 的段落序号，不使用不稳定页码。",
            "- 岁月与生成数据库仍缺失，相关长团候选只能作为限制性材料。",
        ]
    )
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    chunks = load_chunks()
    candidates = extract_candidates(chunks)
    write_jsonl(candidates)
    write_xlsx(candidates)
    write_summary(candidates)
    print(json.dumps({"chunks": len(chunks), "candidates": len(candidates), "jsonl": str(OUT_JSONL), "xlsx": str(OUT_XLSX)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

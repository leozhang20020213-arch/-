from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
EXEC_PACK = ROOT / "_审查执行包_v0.42.1_解压" / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
CHUNKS = ROOT / "review_output" / "01_chapter_chunks.jsonl"
MAPPING = ROOT / "review_output" / "03_rules_mapped_final_b02.jsonl"
MISSING = ROOT / "review_output" / "04_b03_missing_review_draft.jsonl"
OUT_JSONL = ROOT / "review_output" / "05_coverage_draft.jsonl"
OUT_XLSX = ROOT / "review_output" / "05_coverage.xlsx"
OUT_MD = ROOT / "review_output" / "05_coverage_draft.md"
SUMMARY = ROOT / "work_logs" / "b04_coverage_draft_summary.md"


def norm(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def load_coverage_items() -> list[dict[str, object]]:
    wb = load_workbook(EXEC_PACK, data_only=True)
    ws = wb.worksheets[5]
    items = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        items.append({
            "check_id": str(row[0]),
            "object": str(row[1] or ""),
            "scope": str(row[2] or ""),
            "standard": str(row[3] or ""),
            "priority": str(row[8] or ""),
        })
    return items


def load_jsonl(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def evidence_score(item: dict[str, object], chunks: list[dict[str, object]], mapped: list[dict[str, object]], missing: list[dict[str, object]]) -> tuple[int, dict[str, object] | None, str]:
    text = f"{item['object']} {item['scope']} {item['standard']}"
    keys = [k for k in re.split(r"[/、,，;；\s]+", text) if len(k) >= 2]
    best_chunk = None
    best = 0
    for chunk in chunks:
        body = norm(str(chunk.get("chapter_or_heading", "")) + str(chunk.get("text", "")))
        score = sum(2 for k in keys if norm(k) in body)
        if score > best:
            best = score
            best_chunk = chunk
    mapped_hit = sum(1 for row in mapped if any(norm(k) in norm(str(row.get("mapped_rule_id","")) + str(row.get("mapped_rule_name","")) + str(row.get("excerpt",""))) for k in keys))
    missing_hit = sum(1 for row in missing if any(norm(k) in norm(str(row.get("rule_id","")) + str(row.get("rule_name","")) + str(row.get("impact",""))) for k in keys))
    total = best + min(mapped_hit, 5) * 2 - min(missing_hit, 3) * 2
    if missing_hit:
        source = "存在B03缺失/弱证据关联"
    elif mapped_hit and best >= 4:
        source = "正文与映射均有证据"
    elif mapped_hit:
        source = "映射有证据，正文需复核"
    elif best >= 4:
        source = "正文有弱证据，映射需复核"
    else:
        source = "证据不足"
    return total, best_chunk, source


def classify(score: int, source: str) -> tuple[str, str]:
    if "缺失" in source:
        return "需复核", "与B03缺失/弱证据项有关，不能直接判定覆盖。"
    if score >= 10:
        return "基本覆盖-待人工确认", "正文和映射均有较强证据。"
    if score >= 5:
        return "部分覆盖-需补证据", "有证据但不足以确认正文、示例、卡片、数据库均覆盖。"
    return "覆盖不足-待确认", "自动检索证据不足，需要人工审查。"


def make_excerpt(chunk: dict[str, object] | None) -> str:
    if not chunk:
        return ""
    return str(chunk.get("text", ""))[:260].replace("\n", " ")


def build_rows() -> list[dict[str, object]]:
    items = load_coverage_items()
    chunks = load_jsonl(CHUNKS)
    mapped = load_jsonl(MAPPING)
    missing = load_jsonl(MISSING)
    rows = []
    for item in items:
        score, chunk, source = evidence_score(item, chunks, mapped, missing)
        result, note = classify(score, source)
        rows.append({
            **item,
            "draft_result": result,
            "evidence_score": score,
            "evidence_source": source,
            "evidence_chunk_id": "" if chunk is None else chunk.get("chunk_id"),
            "source_file": "" if chunk is None else chunk.get("source_file"),
            "position": "" if chunk is None else json.dumps(chunk.get("position"), ensure_ascii=False),
            "excerpt": make_excerpt(chunk),
            "handling_note": note,
            "limitations": "自动覆盖率草案；只能提示覆盖风险，不直接修改规则书。岁月数据库缺失时长团/年龄/家庭/产业覆盖为限制性。",
        })
    return rows


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "覆盖率审查执行表"
    headers = ["检查ID", "检查对象", "检查范围", "判定标准", "执行结果", "证据分数", "证据来源", "证据位置", "缺失说明", "处理建议", "优先级", "备注", "来源文件"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["check_id"], row["object"], row["scope"], row["standard"], row["draft_result"],
            row["evidence_score"], row["evidence_source"], f"{row['evidence_chunk_id']} {row['position']}",
            row["excerpt"], row["handling_note"], row["priority"], row["limitations"], row["source_file"],
        ])
    fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 30, 34, 46, 22, 10, 24, 24, 70, 50, 10, 54, 48]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["draft_result"]] = counts.get(row["draft_result"], 0) + 1
    lines = ["# B04 覆盖率审查草案", "", "生成时间：2026-06-11 06:33 Asia/Shanghai", "", "## 统计", "", "| 结果 | 数量 |", "| --- | ---: |"]
    for k, v in sorted(counts.items()):
        lines.append(f"| {k} | {v} |")
    lines.extend(["", "## P0覆盖风险", ""])
    for row in rows:
        if row["priority"] == "P0" and row["draft_result"] != "基本覆盖-待人工确认":
            lines.append(f"- `{row['check_id']}` {row['object']}：{row['draft_result']}，{row['handling_note']}")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row["draft_result"]] = counts.get(row["draft_result"], 0) + 1
    lines = ["# B04 覆盖率审查草案摘要", "", "生成时间：2026-06-11 06:33 Asia/Shanghai", "", f"- 覆盖项：{len(rows)}"]
    for k, v in sorted(counts.items()):
        lines.append(f"- {k}：{v}")
    lines.extend([f"- JSONL：`{OUT_JSONL}`", f"- XLSX：`{OUT_XLSX}`", f"- Markdown：`{OUT_MD}`", "", "限制：本轮为自动覆盖率草案，不修改规则书。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_rows()
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"coverage_items": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

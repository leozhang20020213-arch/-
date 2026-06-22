from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
QUEUE = ROOT / "review_output" / "03_p0_manual_review_queue.jsonl"
CHUNKS = ROOT / "review_output" / "01_chapter_chunks.jsonl"
OUT_JSONL = ROOT / "review_output" / "03_p0_manual_queue_triaged.jsonl"
OUT_XLSX = ROOT / "review_output" / "03_p0_manual_queue_triaged.xlsx"
SUMMARY = ROOT / "work_logs" / "b02_manual_queue_triage_summary.md"


EXTRA_HINTS = {
    "SYS-001": ["江湖人生", "职业", "等级", "门派", "背景"],
    "SLOT-004": ["得利", "万能", "兑换", "授权", "菜单", "装备", "关键词"],
    "CBT-001": ["伤害", "结算", "公式", "命中", "兵器"],
    "CBT-005": ["普通敌人", "破绽", "削得利", "退势", "轻反应"],
    "CBT-007": ["反击链", "反击", "不触发", "终止", "Boss"],
    "ACT-001": ["协助", "帮助", "援护"],
    "ACT-002": ["强行尝试", "临场", "尝试", "我能不能"],
    "ACT-004": ["非杀伤", "制服", "不杀", "活捉"],
    "ACT-005": ["环境互动", "环境", "机关", "地形"],
    "FAME-003": ["名望", "调用", "入口", "消耗", "世界权限"],
    "REST-001": ["休整", "三阶段", "疗伤", "修行"],
    "REST-003": ["年龄", "节点", "线性", "惩罚"],
    "FAC-001": ["世界投放", "投放链", "势力", "地图"],
    "CARD-001": ["武器卡", "武器", "卡"],
    "CARD-002": ["招式卡", "招式", "卡"],
    "CARD-003": ["内功卡", "内功", "卡"],
    "REC-001": ["三轨记录", "记录表", "DM", "危机", "解密", "战役得利"],
    "EX-001": ["得利强度表", "得利", "强度", "示例"],
}


def norm(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def load_jsonl(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def score(row: dict[str, object], chunk: dict[str, object]) -> int:
    text = norm(str(chunk.get("chapter_or_heading", "")) + "\n" + str(chunk.get("text", "")))
    rid = str(row["rule_id"])
    value = 0
    for hint in EXTRA_HINTS.get(rid, []):
        h = norm(hint)
        if h and h in text:
            value += 3
    for token in re.split(r"[：:/、，,（）()]+", str(row["rule_name"])):
        t = norm(token)
        if len(t) >= 2 and t in text:
            value += 2
    return value


def excerpt(chunk: dict[str, object], hints: list[str]) -> str:
    text = str(chunk.get("text", ""))
    pos = None
    for hint in hints:
        found = text.find(hint)
        if found >= 0:
            pos = found
            break
    if pos is None:
        return text[:260].replace("\n", " ")
    return text[max(0, pos - 90) : min(len(text), pos + 220)].replace("\n", " ")


def triage(queue: list[dict[str, object]], chunks: list[dict[str, object]]) -> list[dict[str, object]]:
    rows = []
    for item in queue:
        ranked = sorted(((score(item, c), c) for c in chunks), key=lambda pair: (-pair[0], str(pair[1].get("chunk_id"))))
        best_score, best = ranked[0]
        original_status = str(item["review_status"])
        if original_status == "疑似命中-需人工复核" and best_score >= 6:
            triage_status = "确认命中-可补映射"
            action = "下一轮可合并进映射，但仍需在B03引用证据摘录。"
        elif best_score >= 4:
            triage_status = "继续待查-有弱证据"
            action = "进入B03前人工阅读证据块，确认是否真支持Rule_ID。"
        else:
            triage_status = "确认疑似缺失-进B03缺失审查"
            action = "B03中按缺失候选处理；不得直接改规则，先列证据和影响。"
        rows.append(
            {
                "rule_id": item["rule_id"],
                "rule_name": item["rule_name"],
                "original_status": original_status,
                "triage_status": triage_status,
                "triage_action": action,
                "evidence_score": best_score,
                "evidence_chunk_id": best.get("chunk_id"),
                "source_type": best.get("source_type"),
                "source_file": best.get("source_file"),
                "chapter_or_heading": best.get("chapter_or_heading"),
                "position": best.get("position"),
                "excerpt": excerpt(best, EXTRA_HINTS.get(str(item["rule_id"]), [])),
                "limitations": "自动队列分类；B03仍需保留证据和限制说明。岁月数据库缺失时REST/FAC相关结论为限制性。",
            }
        )
    return rows


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "人工队列分类"
    headers = ["Rule_ID", "规则名", "原状态", "分类结果", "建议动作", "证据分数", "证据chunk", "来源", "章节", "位置", "摘录", "来源文件", "限制"]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row["rule_id"],
                row["rule_name"],
                row["original_status"],
                row["triage_status"],
                row["triage_action"],
                row["evidence_score"],
                row["evidence_chunk_id"],
                row["source_type"],
                row["chapter_or_heading"],
                json.dumps(row["position"], ensure_ascii=False),
                row["excerpt"],
                row["source_file"],
                row["limitations"],
            ]
        )
    fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 30, 22, 22, 44, 10, 18, 12, 24, 22, 70, 48, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_summary(rows: list[dict[str, object]]) -> None:
    counts = {}
    for row in rows:
        counts[row["triage_status"]] = counts.get(row["triage_status"], 0) + 1
    lines = [
        "# B02 人工队列分类摘要",
        "",
        "生成时间：2026-06-11 04:03 Asia/Shanghai",
        "",
        "## 输出",
        "",
        f"- JSONL：`{OUT_JSONL}`",
        f"- XLSX：`{OUT_XLSX}`",
        f"- 队列项：{len(rows)}",
        "",
        "## 分类结果",
        "",
        "| 分类 | 数量 |",
        "| --- | ---: |",
    ]
    for k, v in sorted(counts.items()):
        lines.append(f"| {k} | {v} |")
    lines.extend(["", "## 进入B03缺失审查的项目", ""])
    for row in rows:
        if row["triage_status"] == "确认疑似缺失-进B03缺失审查":
            lines.append(f"- `{row['rule_id']}` {row['rule_name']}，证据分数 {row['evidence_score']}")
    lines.extend(["", "## 限制", "", "- 本轮仍不修改规则书。", "- 分类结果用于B03冲突/缺失审查的输入，不是最终修订方案。", "- 岁月数据库仍缺失，REST/FAC相关结论继续限制性处理。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    queue = load_jsonl(QUEUE)
    chunks = load_jsonl(CHUNKS)
    rows = triage(queue, chunks)
    write_jsonl(rows)
    write_xlsx(rows)
    write_summary(rows)
    print(json.dumps({"queue": len(queue), "triaged": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

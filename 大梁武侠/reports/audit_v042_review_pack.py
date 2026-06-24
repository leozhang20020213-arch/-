from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "reports" / "raw_logs"
REPORTS = ROOT / "reports"


def rows_for(ws):
    raw = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(vals):
            raw.append(vals)
    if not raw:
        return [], []
    header_idx = max(range(min(6, len(raw))), key=lambda i: sum(1 for c in raw[i] if c))
    headers = raw[header_idx]
    rows = []
    for vals in raw[header_idx + 1 :]:
        if not any(vals):
            continue
        item = {headers[i] if i < len(headers) and headers[i] else f"字段{i+1}": vals[i] if i < len(vals) else "" for i in range(max(len(headers), len(vals)))}
        rows.append(item)
    return headers, rows


def main() -> None:
    REPORTS.mkdir(exist_ok=True)
    RAW.mkdir(parents=True, exist_ok=True)
    path = next(ROOT.glob("*规则工程化审查包*v0.42.xlsx"))
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        headers, rows = rows_for(ws)
        sheets[ws.title] = {"headers": headers, "rows": rows, "row_count": len(rows)}
    wb.close()

    findings = []
    actions = []

    required_sheets = [
        "README",
        "规则ID命名规范",
        "规则索引_MASTER",
        "术语索引",
        "数据库表索引",
        "冲突审查表",
        "覆盖率审查表",
        "低使用与死规则审查",
        "数值测试矩阵",
        "跑团测试矩阵",
        "v0.42修订任务清单",
        "玩家规则书扩写计划",
        "DM规则书扩写计划",
        "Codex审查流程",
    ]
    missing_sheets = [s for s in required_sheets if s not in sheets]
    if missing_sheets:
        findings.append(("P0", "缺少必要工作表", "、".join(missing_sheets)))
    else:
        findings.append(("OK", "工作表完整", f"包含 {len(required_sheets)} 个预期工作表。"))

    rules = sheets.get("规则索引_MASTER", {}).get("rows", [])
    rule_ids = [r.get("Rule_ID", "") for r in rules if r.get("Rule_ID")]
    dup_rule_ids = [rid for rid, count in Counter(rule_ids).items() if count > 1]
    if dup_rule_ids:
        findings.append(("P0", "Rule_ID 重复", "、".join(dup_rule_ids)))
    else:
        findings.append(("OK", "Rule_ID 无重复", f"规则索引包含 {len(rule_ids)} 个核心 Rule_ID。"))
    p0_rules = [r for r in rules if r.get("优先级") == "P0"]
    must_fix_rules = [r for r in rules if r.get("当前状态") == "必须修"]
    if not must_fix_rules:
        findings.append(("P1", "规则索引缺少必须修标记", "建议至少标出熟练档位、武学容量、普通敌人破绽兑现等必须修。"))
    else:
        findings.append(("OK", "必须修规则已标记", "、".join(f"{r.get('Rule_ID')} {r.get('规则名称')}" for r in must_fix_rules[:8])))

    terms = sheets.get("术语索引", {}).get("rows", [])
    term_ids = [r.get("术语ID", "") for r in terms if r.get("术语ID")]
    dup_term_ids = [tid for tid, count in Counter(term_ids).items() if count > 1]
    undefined_terms = [r for r in terms if not r.get("最终定义")]
    if dup_term_ids:
        findings.append(("P0", "术语ID 重复", "、".join(dup_term_ids)))
    if undefined_terms:
        findings.append(("P1", "术语缺少最终定义", f"{len(undefined_terms)} 条术语缺定义。"))
    else:
        findings.append(("OK", "术语均有最终定义", f"术语索引 {len(terms)} 条。"))

    db_rows = sheets.get("数据库表索引", {}).get("rows", [])
    db_risks = [r for r in db_rows if r.get("缺失风险")]
    if db_rows:
        findings.append(("OK", "数据库表索引存在", f"{len(db_rows)} 个表组，{len(db_risks)} 个已写缺失风险。"))
    else:
        findings.append(("P0", "数据库表索引为空", "无法进行规则-数据库覆盖率审查。"))

    conflicts = sheets.get("冲突审查表", {}).get("rows", [])
    pending_conflicts = [r for r in conflicts if r.get("结果栏") in {"", "待填"}]
    if pending_conflicts:
        findings.append(("P0", "冲突审查尚未执行", f"{len(pending_conflicts)} 项仍为待填。"))
    conflict_topics = "；".join(r.get("冲突主题", "") for r in conflicts[:8])
    findings.append(("INFO", "已覆盖冲突主题", conflict_topics))

    coverage = sheets.get("覆盖率审查表", {}).get("rows", [])
    pending_coverage = [r for r in coverage if r.get("结果栏") in {"", "待填"}]
    if pending_coverage:
        findings.append(("P0", "覆盖率审查尚未执行", f"{len(pending_coverage)} 项仍为待填。"))

    low_use = sheets.get("低使用与死规则审查", {}).get("rows", [])
    low_names = [r.get("规则/内容", "") for r in low_use]
    expected_low = ["战役得利", "协助", "强行尝试", "稳控机会", "非杀伤", "环境互动"]
    missed_low = [x for x in expected_low if not any(x in n for n in low_names)]
    if missed_low:
        findings.append(("P1", "低使用项承接不完整", "缺少：" + "、".join(missed_low)))
    else:
        findings.append(("OK", "低使用项承接前测结果", "战役得利、协助、强行尝试、稳控机会、非杀伤、环境互动均已列入。"))

    math_rows = sheets.get("数值测试矩阵", {}).get("rows", [])
    play_rows = sheets.get("跑团测试矩阵", {}).get("rows", [])
    if len(math_rows) < 10:
        findings.append(("P1", "数值测试矩阵偏少", f"当前 {len(math_rows)} 项；原需求含武学、内功、轻功、武器、防御、爆发、暗器、敌人、资源、休整、岁月等更多测试。"))
    else:
        findings.append(("OK", "数值测试矩阵有基本覆盖", f"{len(math_rows)} 项。"))
    if len(play_rows) < 8:
        findings.append(("P1", "跑团测试矩阵偏少", f"当前 {len(play_rows)} 项；压力场景需要继续展开到 20 项。"))
    else:
        findings.append(("OK", "跑团测试矩阵有基本覆盖", f"{len(play_rows)} 项。"))

    tasks = sheets.get("v0.42修订任务清单", {}).get("rows", [])
    task_ids = [r.get("任务ID", "") for r in tasks if r.get("任务ID")]
    dup_task_ids = [tid for tid, count in Counter(task_ids).items() if count > 1]
    if dup_task_ids:
        findings.append(("P0", "任务ID重复", "、".join(dup_task_ids)))
    p0_tasks = [r for r in tasks if r.get("优先级") == "P0"]
    findings.append(("INFO", "P0修订任务", "；".join(f"{r.get('任务ID')} {r.get('修订任务')}" for r in p0_tasks[:10])))

    player_plan = sheets.get("玩家规则书扩写计划", {}).get("rows", [])
    dm_plan = sheets.get("DM规则书扩写计划", {}).get("rows", [])
    player_pages = sum(int(float(r.get("目标页数", "0") or 0)) for r in player_plan if str(r.get("目标页数", "")).replace(".", "", 1).isdigit())
    dm_pages = sum(int(float(r.get("目标页数", "0") or 0)) for r in dm_plan if str(r.get("目标页数", "")).replace(".", "", 1).isdigit())
    findings.append(("INFO", "扩写页数规划", f"玩家书已列章节目标约 {player_pages} 页；DM书已列章节目标约 {dm_pages} 页。"))
    if player_pages < 70:
        findings.append(("P1", "玩家书扩写计划页数不足", "README目标约80~100页，当前列出的章节合计偏少，可能还有后续章节未列全。"))
    if dm_pages < 110:
        findings.append(("P1", "DM书扩写计划页数不足", "README目标约120~160页，当前列出的章节合计偏少，可能还有后续章节未列全。"))

    flow_rows = sheets.get("Codex审查流程", {}).get("rows", [])
    flow_steps = [r.get("步骤", "") for r in flow_rows]
    if len(flow_rows) < 9:
        findings.append(("P1", "Codex审查流程不完整", f"当前 {len(flow_rows)} 步，README预期一次性完整审阅应覆盖更多输出。"))
    else:
        findings.append(("OK", "Codex审查流程存在", f"{len(flow_rows)} 步。"))

    actions.extend([
        ("P0", "先执行冲突审查表", "把 CON001-CON006 等待填项跑成结果，尤其熟练档位、武学容量、得利授权、Aim、架势、情景资源。"),
        ("P0", "先执行覆盖率审查表", "把 COV001-COV006 等待填项跑成结果，确认 Rule_ID 是否真的进入玩家书/DM书正文。"),
        ("P0", "扩充规则索引_MASTER", "当前 54 条是核心索引，不足以覆盖完整玩家书/DM书；应接入自动抽取的 137 条候选规则并映射 Rule_ID。"),
        ("P0", "补齐页数计划", "玩家书扩写计划和DM书扩写计划当前页数偏少，应补全后续章节。"),
        ("P1", "扩充数学测试矩阵", "补武学成长、轻功、敌人60个、休整50行动、岁月20/30/50年等测试项。"),
        ("P1", "扩充跑团压力矩阵", "补齐20类极限压力跑团。"),
        ("P1", "给任务清单增加状态字段", "建议增加负责人、输入文件、输出文件、验收结果、阻塞项。"),
    ])

    payload = {
        "file": path.name,
        "sheets": {k: {"row_count": v["row_count"], "headers": v["headers"]} for k, v in sheets.items()},
        "findings": findings,
        "actions": actions,
    }
    (RAW / "v042_review_pack_audit.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# 大梁武侠TRPG v0.42 规则工程化审查包审查报告",
        "",
        "## 审查对象",
        "",
        f"- 文件：`{path.name}`",
        "- 类型：Excel 工程化审查包，不是压缩包。",
        "- 审查目标：判断该审查包是否足以支撑后续 Codex 全量审查、规则书扩写、数据库覆盖率检查、数值测试和跑团测试。",
        "",
        "## 总体结论",
        "",
        "v0.42 审查包方向正确，已经把前期测试暴露的关键问题工程化为 Rule_ID、冲突审查、覆盖率审查、低使用项、测试矩阵和修订任务。",
        "",
        "但它目前更像“审查蓝图/任务规划表”，还不是可直接宣称完成审查的结果包。核心原因是多个审查表的结果栏仍为待填，规则索引_MASTER 只有核心 Rule_ID，尚未映射完整规则书全文，数学测试和跑团压力矩阵也未覆盖原需求的全部测试项目。",
        "",
        "## 工作表概览",
        "",
        "| 工作表 | 数据行数 | 说明 |",
        "| --- | ---: | --- |",
    ]
    for name, meta in sheets.items():
        lines.append(f"| {name} | {meta['row_count']} | 字段：{'、'.join(meta['headers'][:6])} |")
    lines += [
        "",
        "## 审查发现",
        "",
        "| 等级 | 项目 | 说明 |",
        "| --- | --- | --- |",
    ]
    for level, item, detail in findings:
        lines.append(f"| {level} | {item} | {detail} |")
    lines += [
        "",
        "## 与前期测试结论的承接情况",
        "",
        "### 承接良好",
        "",
        "- 普通敌人频繁反击问题已进入 `v0.42修订任务清单`：`T042-003 修普通敌人破绽兑现`。",
        "- 熟练档位冲突已进入 `T042-001`。",
        "- 武学容量冲突已进入 `T042-002`。",
        "- 高代价成功话术已进入 `T042-004`。",
        "- 得利1/2/3强度表已进入 `T042-005`。",
        "- 玩家行动菜单已进入 `T042-006`。",
        "- 低使用项中，战役得利、协助、强行尝试、稳控机会、非杀伤、环境互动都已被列出。",
        "",
        "### 仍需补强",
        "",
        "- `武学谱修行`、`内功成长`、`名望/人情/情报/把柄消费入口`、`家庭/弟子/产业长期权限` 在任务清单中的显式程度仍不够。",
        "- `岁月与生成数据库` 的缺失问题没有在审查包中作为输入风险突出显示。",
        "- 数学测试矩阵尚未覆盖所有原始要求，如 60 个敌人逐一测试、50 个休整行动、20/30/50 年岁月长期测试等。",
        "- 跑团测试矩阵尚未展开到 20 项极限压力跑团。",
        "",
        "## P0 问题",
        "",
        "1. `冲突审查表` 和 `覆盖率审查表` 多数结果栏仍为待填。该包目前不能替代实际审查结果。",
        "2. `规则索引_MASTER` 只有核心 Rule_ID，尚未覆盖两本规则书全文。需要把自动抽取规则候选映射进去。",
        "3. 扩写计划页数合计不足 README 中的目标页数，说明玩家书/DM书扩写目录仍未完整展开。",
        "4. 缺少对输入文件缺口的显式风险表，尤其是岁月数据库缺失会影响长团和年龄系统审查。",
        "",
        "## 建议处理动作",
        "",
        "| 优先级 | 动作 | 说明 |",
        "| --- | --- | --- |",
    ]
    for level, action, detail in actions:
        lines.append(f"| {level} | {action} | {detail} |")
    lines += [
        "",
        "## 可直接用于下一步 Codex 审查的顺序",
        "",
        "1. 用 `规则ID命名规范` 固定 Rule_ID 前缀，不再临时命名。",
        "2. 将玩家书和DM书全文切块，抽取规则候选。",
        "3. 把候选规则映射到 `规则索引_MASTER`，新增缺失 Rule_ID。",
        "4. 执行 `冲突审查表`，把结果栏从待填变成“通过/冲突/缺失/需改”。",
        "5. 执行 `覆盖率审查表`，检查每个 Rule_ID 是否有正文、示例、卡片、数据库引用。",
        "6. 按 `v0.42修订任务清单` 优先完成 P0 任务。",
        "7. 扩展 `数值测试矩阵` 和 `跑团测试矩阵` 到完整需求。",
        "8. 输出 v0.42 审查结果包，而不是只保留审查计划包。",
        "",
        "## 最终判断",
        "",
        "该 v0.42 审查包值得作为下一轮规则工程化审查的主控文件。它已经正确把“普通敌人反击、熟练档位、武学容量、高代价成功、得利强度、行动菜单”等前测核心问题转成可执行任务。",
        "",
        "但在实际审查前，必须补齐结果栏、扩展规则索引、补全测试矩阵，并明确输入文件缺口。否则它只能说明“准备审查什么”，不能说明“已经审查通过”。",
    ]
    (REPORTS / "v0.42规则工程化审查包_审查报告.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()

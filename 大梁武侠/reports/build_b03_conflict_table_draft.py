from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(r"D:\trpg\大梁武侠")
EXEC_PACK = ROOT / "_审查执行包_v0.42.1_解压" / "大梁武侠TRPG_规则工程化审查执行包_v0.42.1.xlsx"
CHUNKS = ROOT / "review_output" / "01_chapter_chunks.jsonl"
OUT_JSONL = ROOT / "review_output" / "04_conflicts_draft.jsonl"
OUT_XLSX = ROOT / "review_output" / "04_conflicts.xlsx"
OUT_MD = ROOT / "review_output" / "04_conflicts_draft.md"
SUMMARY = ROOT / "work_logs" / "b03_conflict_table_draft_summary.md"


def norm(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def load_conflicts() -> list[dict[str, object]]:
    wb = load_workbook(EXEC_PACK, data_only=True)
    ws = wb.worksheets[4]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        rows.append(
            {
                "conflict_id": str(row[0]),
                "topic": str(row[1] or ""),
                "keywords": str(row[2] or ""),
                "final_position": str(row[3] or ""),
                "priority": str(row[9] or ""),
            }
        )
    return rows


def load_chunks() -> list[dict[str, object]]:
    with CHUNKS.open("r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]


def split_keywords(value: str) -> list[str]:
    parts = re.split(r"[/、,，;；\s]+", value)
    return [p.strip() for p in parts if len(p.strip()) >= 2]


def score(conflict: dict[str, object], chunk: dict[str, object]) -> int:
    text = norm(str(chunk.get("chapter_or_heading", "")) + "\n" + str(chunk.get("text", "")))
    value = 0
    for key in split_keywords(str(conflict["keywords"])) + split_keywords(str(conflict["topic"])):
        if norm(key) in text:
            value += 2
    for key in split_keywords(str(conflict["final_position"])):
        if norm(key) in text:
            value += 1
    return value


def excerpt(chunk: dict[str, object], keywords: list[str]) -> str:
    text = str(chunk.get("text", ""))
    pos = None
    for key in keywords:
        found = text.find(key)
        if found >= 0:
            pos = found
            break
    if pos is None:
        return text[:260].replace("\n", " ")
    return text[max(0, pos - 100) : min(len(text), pos + 240)].replace("\n", " ")


def classify(conflict: dict[str, object], best_score: int, second_score: int) -> tuple[str, str]:
    if best_score >= 8 and second_score >= 5:
        return "需人工判断", "找到多处相关证据，可能存在多口径或跨章节分散，需要人工比对原文。"
    if best_score >= 6:
        return "有证据-待复核", "找到较强相关证据，但本轮不直接判定通过或冲突。"
    if best_score >= 3:
        return "弱证据-待复核", "只有弱相关证据，需要人工阅读证据块。"
    return "缺失-待确认", "未在章节块中找到足够相关证据，需人工全文检索确认。"


def build_rows(conflicts: list[dict[str, object]], chunks: list[dict[str, object]]) -> list[dict[str, object]]:
    rows = []
    for conflict in conflicts:
        ranked = sorted(((score(conflict, chunk), chunk) for chunk in chunks), key=lambda pair: (-pair[0], str(pair[1].get("chunk_id"))))
        best_score, best = ranked[0]
        second_score = ranked[1][0] if len(ranked) > 1 else 0
        result, note = classify(conflict, best_score, second_score)
        keys = split_keywords(str(conflict["keywords"])) + split_keywords(str(conflict["topic"]))
        rows.append(
            {
                **conflict,
                "draft_result": result,
                "evidence_score": best_score,
                "second_evidence_score": second_score,
                "evidence_chunk_id": best.get("chunk_id"),
                "source_type": best.get("source_type"),
                "source_file": best.get("source_file"),
                "chapter_or_heading": best.get("chapter_or_heading"),
                "position": best.get("position"),
                "excerpt": excerpt(best, keys),
                "handling_note": note,
                "limitations": "自动冲突草案；需人工阅读证据后再判定通过/冲突/缺失/需改。岁月数据库缺失时长团相关结论限制性处理。",
            }
        )
    return rows


def write_jsonl(rows: list[dict[str, object]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "冲突审查执行表"
    headers = ["冲突ID", "冲突主题", "检查关键词", "最终口径", "执行结果", "证据分数", "次证据分数", "证据位置", "冲突原文摘录", "处理建议", "需修改Rule_ID", "优先级", "备注", "来源文件"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["conflict_id"],
            row["topic"],
            row["keywords"],
            row["final_position"],
            row["draft_result"],
            row["evidence_score"],
            row["second_evidence_score"],
            f"{row['evidence_chunk_id']} {json.dumps(row['position'], ensure_ascii=False)}",
            row["excerpt"],
            row["handling_note"],
            "待B03人工确认",
            row["priority"],
            row["limitations"],
            row["source_file"],
        ])
    fill = PatternFill("solid", fgColor="F8CBAD")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 24, 42, 42, 18, 10, 10, 24, 70, 52, 20, 10, 52, 48]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for data_row in ws.iter_rows(min_row=2):
        for cell in data_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(rows: list[dict[str, object]]) -> None:
    counts = {}
    for row in rows:
        counts[row["draft_result"]] = counts.get(row["draft_result"], 0) + 1
    lines = ["# B03 冲突审查草案", "", "生成时间：2026-06-11 05:33 Asia/Shanghai", "", "## 统计", "", "| 草案结果 | 数量 |", "| --- | ---: |"]
    for key, value in sorted(counts.items()):
        lines.append(f"| {key} | {value} |")
    lines.extend(["", "## P0重点项", ""])
    for row in rows:
        if row["priority"] == "P0":
            lines.extend([
                f"### `{row['conflict_id']}` {row['topic']}",
                f"- 草案结果：{row['draft_result']}",
                f"- 证据：`{row['evidence_chunk_id']}`，分数 {row['evidence_score']}",
                f"- 处理建议：{row['handling_note']}",
                "",
            ])
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[dict[str, object]]) -> None:
    counts = {}
    for row in rows:
        counts[row["draft_result"]] = counts.get(row["draft_result"], 0) + 1
    lines = ["# B03 冲突审查草案摘要", "", "生成时间：2026-06-11 05:33 Asia/Shanghai", "", f"- 冲突项：{len(rows)}"]
    for key, value in sorted(counts.items()):
        lines.append(f"- {key}：{value}")
    lines.extend([f"- JSONL：`{OUT_JSONL}`", f"- XLSX：`{OUT_XLSX}`", f"- Markdown：`{OUT_MD}`", "", "限制：本轮为自动草案，不直接判定最终冲突，也不修改规则书。"])
    SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    conflicts = load_conflicts()
    chunks = load_chunks()
    rows = build_rows(conflicts, chunks)
    write_jsonl(rows)
    write_xlsx(rows)
    write_md(rows)
    write_summary(rows)
    print(json.dumps({"conflicts": len(conflicts), "rows": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
